import logging
import sys
import subprocess
import csv
import os

# Configuración de logging
logging.basicConfig(filename="registros/script.log", level=logging.INFO, encoding="utf-8",
                    format="%(asctime)s - %(levelname)s - %(module)s:%(lineno)d - %(message)s")

# verificar e instalar openpyxl y pandas si no están instalados
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Alignment, numbers
    from openpyxl.utils.dataframe import dataframe_to_rows
    logging.info("openpyxl ya está instalado.")
except ImportError:
    print("openpyxl no está instalado. Instalando...")
    logging.info("openpyxl no está instalado. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Alignment, numbers
    from openpyxl.utils.dataframe import dataframe_to_rows
try:
    import pandas as pd
    logging.info("pandas ya está instalado.")
except ImportError:
    print("pandas no está instalado. Instalando...")
    logging.info("pandas no está instalado. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
    import pandas as pd


def get_excel_files(directory):
    return [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith(('.xlsx', '.csv'))]

def read_csv_to_dataframe(csv_file):
    with open(csv_file, 'r') as file:
        try:
            df = pd.read_csv(file)
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
    return df

def auto_adjust_column_width(sheet):
    for col in sheet.columns:
        # Filtra las celdas con valores no nulos
        non_empty_cells = [cell for cell in col if cell.value]
        
        # Solo ajusta el ancho si hay celdas con valores
        if non_empty_cells:
            max_length = max(len(str(cell.value)) for cell in non_empty_cells)
            adjusted_width = max_length + 2
            sheet.column_dimensions[col[0].column_letter].width = adjusted_width

def copy_cell_styles(src_cell, dest_cell):
    """Copia los estilos de una celda de origen a una celda de destino."""
    if src_cell.has_style:
        dest_cell.font = src_cell.font.copy()
        # Copiar cada lado del borde individualmente
        dest_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom,
            diagonal=src_cell.border.diagonal,
            diagonal_direction=src_cell.border.diagonal_direction,
            outline=src_cell.border.outline,
            vertical=src_cell.border.vertical,
            horizontal=src_cell.border.horizontal
        )
        dest_cell.fill = src_cell.fill.copy()
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = src_cell.protection.copy()
        dest_cell.alignment = src_cell.alignment.copy()

def process_excel_files(directory, master_file_name):
    master_workbook = Workbook()
    master_workbook.remove(master_workbook.active)  # Remove the default sheet

    excel_files = get_excel_files(directory)
    
    for file in excel_files:
        logging.info(f"Processing file: {file}")
        file_name = os.path.splitext(os.path.basename(file))[0]
        
        if file.endswith('.xlsx'):
            workbook = load_workbook(file)
            sheet_number = 1
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                if ws.max_row > 1:
                    new_sheet_name = f"{file_name}-{sheet_number}" if sheet_number > 1 else file_name
                    master_sheet = master_workbook.create_sheet(title=new_sheet_name)

                    # Copia de datos (vieja)
                    # for row in ws.iter_rows(values_only=True):
                    #     master_sheet.append(row)

                    # Copia de datos y estilos
                    for row in ws.iter_rows():
                        master_row = []
                        for cell in row:
                            new_cell = master_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                            copy_cell_styles(cell, new_cell)
                        master_sheet.append(master_row)
                        
                    auto_adjust_column_width(master_sheet)
                    logging.info(f"Added sheet {new_sheet_name} from {file}")
                    sheet_number += 1
                    
                else:
                    logging.info(f"Skipped empty sheet {sheet} in {file}")

        elif file.endswith('.csv'):
            df = read_csv_to_dataframe(file)
            if not df.empty:
                master_sheet = master_workbook.create_sheet(title=file_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    master_sheet.append(r)
                auto_adjust_column_width(master_sheet)
                logging.info(f"Added sheet {file_name} from {file}")
            else:
                logging.info(f"Skipped empty CSV file {file}")

    master_file_path = os.path.join(directory, master_file_name)
    master_workbook.save(master_file_path)
    logging.info(f"Master Excel file '{master_file_path}' created successfully.")

def main():
    # Registrar un nuevo lanzamiento del script
    logging.info("\n\n-------NEW MASTER EXCEL-------\n")

    directory = "C:\\Linux"
    master_file_name = input("Enter the name for the master Excel file: ")
    process_excel_files(directory, master_file_name + ".xlsx")

    print("\nMasterExcel finalizado\n")

if __name__ == "__main__":
    main()
