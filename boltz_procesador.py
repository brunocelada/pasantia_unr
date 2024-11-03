import os
import glob
import logging 
import subprocess
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
import boltz_2_ordenado_SCF

# Configuración de logging
logging.basicConfig(filename="registros/script.log", level=logging.INFO, encoding="utf-8",
                    format="%(asctime)s - %(levelname)s - %(module)s:%(lineno)d - %(message)s")

# Verificar e instalar openpyxl si no está instalado
try:
    import xlwings as xw
    logging.info("xlwings ya está instalado.")
except ImportError:
    print("xlwings no está instalado. Instalando...")
    logging.info("xlwings no está instalado. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlwings"])
    import xlwings as xw

def calc_percent(ws, value1, value2):
    '''Recordar que trabaja con hasta 1000 valores de boltzmann'''
    # Programar los valores a comparar (en este caso, value1 vs value2)
    ws["F1"] = "Valores a comparar"
    ws["F2"] = value1 # estos valores pueden cambiarse para elegir la comparación entre
    ws["G2"] = value2 # otros tipos de compuestos / isómeros

    # Calcular los porcentajes de la primera etapa
    ws["F6"] = '="% "&F2'
    ws["G6"] = '="% "&G2'
    ws["F7"] = '=SUMIFS(D2:D1000, J2:J1000, F2)'
    ws["G7"] = '=SUMIFS(D2:D1000, J2:J1000, G2)'

    # Mostrar el porcentaje ganador de la comparación
    ws["F9"] = '=IF(F7>G7,F2,G2)'
    ws["F11"] = '=MAX(F7:G7)'

    for i in range(2, 1000):
        ws[f"J{i}"] = f'=IF(A{i}<>"", IF(ISNUMBER(SEARCH($F$2, A{i})), $F$2, IF(ISNUMBER(SEARCH($G$2, A{i})), $G$2, "ERROR!")), "")'
                      # =SI(A2<>""; SI(ESNUMERO(HALLAR($F$2; A2)); $F$2; SI(ESNUMERO(HALLAR($G$2; A2)); $G$2; "error!!")); "")
    
    style_sheet(ws)

def style_sheet(sheet):
    # Definir las celdas a las que se le aplicaran cierto estilo
    subtitle_cells = [sheet["F1"], sheet["F6"], sheet["G6"], sheet["F9"], sheet["F11"]]
    centered_cells = [sheet["F1"], sheet["F2"], sheet["G2"], sheet["F6"], sheet["F7"], 
                      sheet["G6"], sheet["G7"], sheet["F9"], sheet["F11"]]
    gray_color = []
    for i in range(2, 1000):
        gray_color.append(sheet[f"J{i}"])
    number_2decimal = [sheet["F7"], sheet["G7"], sheet["F11"]]

    # Unificar las celdas previo al aplicado de un estilo
    sheet.merge_cells("F1:G1")
    sheet.merge_cells("F9:G10")
    sheet.merge_cells("F11:G12")

    # Iterar por cada celda aplicando el estilo
    for cell in subtitle_cells:     # Estilo negrita
        cell.font = Font(bold=True, size=12)
    for cell in centered_cells:     # Texto centrado
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in gray_color:
        cell.font = Font(color="808080")
    for cell in number_2decimal:    # 2 Decimales a los números
        cell.number_format = numbers.FORMAT_NUMBER_00

def obtain_looser_files(carpeta, winner):
    archivos = os.listdir(carpeta)
    looser_files = []
    for file in archivos:
        if winner not in file:
            looser_files.append(file)

    return looser_files

def procesar_boltz_excel(carpeta, sheetName, excelName, value1, value2):
        os.chdir(carpeta)
        logging.info(f"Processing file: {excelName}")
        
        try:
            workbook = load_workbook(excelName)
            ws = workbook[sheetName]
                
            # Calcular porcentajes de "endo" vs "exo", y luego "R" vs "S"
            calc_percent(ws, value1, value2)

            logging.info(f"Excel {excelName} modificado")

            workbook.save(excelName)

            # Utiliza otra librería para extraer el valor calculado
            wb = xw.Book(excelName)
            sheet = wb.sheets[sheetName]
            winner = sheet.range("F9").value  # Esto devuelve el valor calculado, no la fórmula
            wb.close()

            # Hacer una lista de los archivos .log que no contienen a winner en su nombre
            looser_files = obtain_looser_files(carpeta, winner)

            return looser_files, winner

        except Exception as e:
            logging.error(f"Error procesando archivo {excelName}: {e}")



def main():
    carpeta_base = "C:\\Linux"

    fValue1 = "endo"
    fValue2 = "exo"
    lValue1 = "-R-"
    lValue2 = "-S-"

    # Ejecutar el script "boltz_2_ordenado_SFC.py" para crear el archivo Boltzmann.xlsx
    boltz_2_ordenado_SCF.main()

    # Carpetas que no se procesarán
    patrones_ignorados = ["Frecuencias Negativas", "Fre", "Relanzar", "Rel", "Termino Mal", "Ter"]

    # Registrar un nuevo lanzamiento del script "boltz_procesador.py"
    logging.info("\n\n----NEW BOLTZ_PROCESADOR----\n")

    carpetas = glob.glob(os.path.join(carpeta_base, "*"))
    
    for carpeta in carpetas:
        nombre_carpeta = os.path.basename(carpeta)

        if nombre_carpeta in patrones_ignorados:
            continue

        # Obtiene el ganador entre fValue1 y fValue2, añadiendo a la lista aquellos archivos .log que no pertenezcan al ganador
        looser_files, winner = procesar_boltz_excel(carpeta, "Boltzmann Data", "Boltzmann.xlsx", fValue1, fValue2)

        # Crea un nuevo excel de Boltzmann para el ganador, y compara entre lValue1 y lValue2.
        boltz_2_ordenado_SCF.read_log_files(carpeta, sheettitle=f"Boltzmann-{winner}", filetitle=f"Boltzmann-{winner}.xlsx", files_ignorados=looser_files)
        looser_files2, winner2 = procesar_boltz_excel(carpeta, f"Boltzmann-{winner}", f"Boltzmann-{winner}.xlsx", lValue1, lValue2)

    print("\nProcesador de Boltzmann Finalizado\n")

if __name__ == "__main__":
    main()
