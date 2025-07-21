import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

"""
----------------------------------------------------------
Análisis de Impacto Marginal con parsing automático
----------------------------------------------------------

- Extrae Normalization, Reduction y Model de la columna 'Modelo'
- Calcula el impacto promedio para cada parte y métrica
- Guarda todas las combinaciones en un solo Excel
"""

# ----------------------------------------
# Configuración básica del logging
# ----------------------------------------
def setup_logging():
    """
    Configura el logging para registrar mensajes en consola con nivel INFO.
    """
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ----------------------------------------
# Función genérica: prompt_input
# Descripción: Solicita al usuario un valor con validación y valor por defecto.
# ----------------------------------------
def prompt_input(message, default=None, validator=None, error_msg="Entrada inválida"):
    """
    Pide al usuario un input con un mensaje personalizado.
    - message: texto que se muestra al usuario.
    - default: valor por defecto si el usuario no ingresa nada.
    - validator: función que recibe el input y retorna True si es válido.
    - error_msg: mensaje de error al reintentar.
    """
    while True:
        try:
            prompt_msg = message
            if default is not None:
                prompt_msg += f" (por defecto '{default}')"
            prompt_msg += ": "
            user_input = input(prompt_msg).strip()
            # Aplica valor por defecto
            result = user_input or default
            # Valida si se proporcionó validador
            if validator and not validator(result):
                raise ValueError(error_msg)
            logger.info(f"Usuario ingresó: {result}")
            return result
        except Exception as e:
            logger.warning(f"{e}. Por favor, intente de nuevo.")

# ----------------------------------------
# Función: prompt_excel_path (usa prompt_input)
# Descripción: Solicita al usuario la ruta del archivo Excel de manera genérica.
# ----------------------------------------
def prompt_excel_path():
    """
    Pide al usuario la ruta del archivo Excel.
    Utiliza prompt_input para manejar validación y valor por defecto.
    """
    return prompt_input(
        message="Ruta del archivo Excel",
        default='C:\\Linux\\Master_Table.xlsx',
        validator=lambda p: bool(p and os.path.splitext(p)[1] in ['.xlsx', '.xls']),
        error_msg="Debe ingresar un nombre de archivo válido con extensión .xlsx o .xls"
    )

def load_master_table(filepath: str, sheet_name: str = None) -> pd.DataFrame:
    """
    Carga el archivo Master_Table.
    """
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    return df

def parse_model_column(df: pd.DataFrame, model_col: str = "Modelo") -> pd.DataFrame:
    """
    Crea columnas nuevas 'Normalization', 'Reduction', 'Model' a partir de 'Modelo'.
    """
    split_cols = df[model_col].str.split('_', expand=True)
    df['Normalization'] = split_cols[0]
    df['Reduction'] = split_cols[1]
    df['Model'] = split_cols[2]
    return df

def compute_marginal_impact(df: pd.DataFrame, part: str, metric: str, descending: bool) -> pd.DataFrame:
    """
    Calcula el impacto marginal promedio para cada categoría de 'part'.
    """
    impacts = []
    groups = df.groupby(part)

    for name, group in groups:
        mean_metric = group[metric].mean()
        impacts.append({
            part: name,
            f'Avg_{metric}': mean_metric,
            'N_models': len(group)
        })

    impact_df = pd.DataFrame(impacts).sort_values(f'Avg_{metric}', ascending=not descending)
    return impact_df

def is_r2(metric: str) -> bool:
    """
    Determina si la métrica es de tipo R2.
    """
    return "R2" in metric

def clean_percentage_columns(df: pd.DataFrame, percentage_columns: list) -> pd.DataFrame:
    """
    Quita '%' y convierte las columnas a float.
    """
    for col in percentage_columns:
        if df[col].dtype == object:
            df[col] = df[col].str.rstrip('%').astype(float)
    return df

# ----------------------------------------
# Función: ajustar_ancho_columnas_excel
# ----------------------------------------
def ajustar_ancho_columnas_excel(path_excel):
    """
    Ajusta el ancho de las columnas automáticamente para cada hoja del archivo Excel.
    """
    wb = load_workbook(path_excel, data_only=True)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(path_excel)
    logger.info(f"Ajuste de anchos de columnas aplicado: {path_excel}")


def main():
    # === CONFIG ===
    filepath = prompt_excel_path()
    sheet_name = "Resumen"
    model_col = "Modelo"

    metrics = [
        "R2_LOOCV",
        "R2_5FoldCV",
        "R2_ObsVsPred",
        "MAE",
        "RMSE",
        "MAPE",
        "Outlier_∆∆G-0.3",
        "Outlier_%ee-30%"
    ]

    # === Cargar y parsear ===
    df = load_master_table(filepath, sheet_name)
    df = parse_model_column(df, model_col)

    # === Limpiar columnas porcentuales ===
    percentage_columns = ["MAPE"]
    df = clean_percentage_columns(df, percentage_columns)

    parts = ["Normalization", "Reduction", "Model"]

    base_dir = os.path.dirname(filepath)
    output_file = os.path.join(base_dir, "Impact_Marginal_AllParts.xlsx")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for part in parts:
            for metric in metrics:
                descending = is_r2(metric)
                impact_df = compute_marginal_impact(df, part, metric, descending)
                sheet_name_excel = f"{part}_{metric}"[:31]
                impact_df.to_excel(writer, sheet_name=sheet_name_excel, index=False)
                print(f"[{part}] Impacto marginal para {metric} agregado. Orden: {'desc' if descending else 'asc'}")

    print(f"\n✅ Archivo generado: {output_file}")

    # Ajusta automáticamente el ancho de las columnas del Excel generado.
    ajustar_ancho_columnas_excel(output_file)

if __name__ == "__main__":
    main()