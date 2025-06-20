# -----------------------------------------------------------------------------------------------------------------------
# Se debe usar este script por el momento en WINDOWS solamente, y tener instalado EXCEL y las librerías necesarias.
# Durante su ejecución, no abrir ni cerrar Excel, ya que podrían ocurrir errores en el script.

# Si aparece el siguiente error:
#     'C:\Users\bruno\AppData\Local\Programs\Python\Python313\Lib\site-packages\openpyxl\worksheet\_reader.py:329: UserWarning: Unknown extension is not supported and will be removed
#   warn(msg)'
# es porque el archivo Excel tiene una extensión o un formato en alguna celda que es desconocido, y openpyxl no puede manejarlo. Esto es normal y no afecta al funcionamiento del script.

# Librerías necesarias:
#     - openpyxl
#     - xlwings

# ------------------------------------------------------------------------------------------------------------------------

import os
import logging
import re
import time
import winsound
import json

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
import xlwings as xw

from last_lines_selector import obtener_listas_de_modelos, pedir_lineas_criterio


# -------------------- CONFIGURACIÓN DE LOGGING --------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)


# -------------------- FUNCIÓN 1 --------------------
def get_column_from_excels_in_subfolders(base_folder, file_prefix, column_letter):
    column_data = {}
    col_idx = column_index_from_string(column_letter)

    for subfolder in os.listdir(base_folder):
        if subfolder.lower().startswith("error") or subfolder.lower().startswith("logs"):
            continue
        subfolder_path = os.path.join(base_folder, subfolder)
        if os.path.isdir(subfolder_path):
            for filename in os.listdir(subfolder_path):
                if filename.startswith(file_prefix) and filename.endswith(".xlsx") and "training" not in filename.lower():
                    path = os.path.join(subfolder_path, filename)
                    wb = load_workbook(path, data_only=True)
                    ws = wb.active
                    values = []
                    row = 1
                    while True:
                        cell_value = ws.cell(row=row, column=col_idx).value
                        if cell_value is None:
                            break
                        values.append(cell_value)
                        row += 1
                    column_data[filename] = (values, subfolder)
                    logging.info(f"Extraídos {len(values)} valores de '{filename}' en columna '{column_letter}'.")
    return column_data

# -------------------- NUEVA FUNCION: Templates + Outliers + Hoja Extra --------------------
def generate_templates_and_calculate_outliers(template_path, data_dict, target_column_letter, output_folder,
                                              l_cell, w_cell, k_cell, v_cell, guardar_templates):
    target_col_idx = column_index_from_string(target_column_letter)
    

    # Limpiar archivos existentes antes de iniciar
    ruta_ddg = os.path.join(os.path.dirname(output_folder), "outlier_dict_ddg.txt")
    ruta_ee = os.path.join(os.path.dirname(output_folder), "outlier_dict_ee.txt")
    open(ruta_ddg, 'w').close()
    open(ruta_ee, 'w').close()

    for model_name, (values, subfolder) in data_dict.items():
        if guardar_templates:
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, f"Template_{model_name}.xlsx")

        else:
            output_path = template_path  # Se usará directamente

        # Copiar valores al template
        wb = load_workbook(template_path)
        ws = wb.active
        for i, val in enumerate(values, start=1):
            ws.cell(row=i, column=target_col_idx, value=val)

        # Si se guarda template, lo guardamos como nuevo archivo
        if guardar_templates:
            wb.save(output_path)
        else:
            wb.save(template_path)  # Sobrescribe temporalmente el original

        wb.close()


        # Calcular outliers con xlwings
        logging.info(f"Calculando outliers para: {model_name}")
        app = None
        book = None
        try:
            app = xw.App(visible=False) # Cambiar a True momentáneamente si se quiere ver cómo cambia el Excel.
            book = app.books.open(output_path)
            sheet = book.sheets[0]

            outliers_ddg = []
            outliers_ee = []
            for i in range(50):
                out_val = round((i + 1) * 0.01, 2)
                
                # logging.info(f"Calculando los outliers con un cuf_off de: {out_val}")
                print(f"Calculando los outliers con un cut_off de: {out_val:.2f}", end='\r')

                sheet.range(l_cell).value = out_val
                sheet.range(w_cell).value = out_val
                book.app.calculate()

                time.sleep(0.2)
                outliers_ddg.append(sheet.range(k_cell).value)
                outliers_ee.append(sheet.range(v_cell).value)

            # Guardar resultados en archivos txt
            with open(ruta_ddg, 'a', encoding='utf-8') as f_ddg:
                f_ddg.write(json.dumps({model_name: outliers_ddg}, ensure_ascii=False) + '\n')
            logging.info(f"Se añadieron los valores de {model_name} al diccionario Outliers_∆∆G.")
            with open(ruta_ee, 'a', encoding='utf-8') as f_ee:
                f_ee.write(json.dumps({model_name: outliers_ee}, ensure_ascii=False) + '\n')
            logging.info(f"Se añadieron los valores de {model_name} al diccionario Outliers_%ee.")

        except Exception as e:
            logging.error(f"Error al procesar {template_path} con xlwings: {e}", exc_info=True)
            continue

        finally:
            if book is not None:
                try:
                    book.close()
                except Exception as e:
                    logging.warning(f"No se pudo cerrar el libro: {e}")
            if app is not None:
                try:
                    app.quit()
                except Exception as e:
                    logging.warning(f"No se pudo cerrar la app de Excel: {e}")

        # Si se guardan templates, agregamos la hoja "Outliers"
        if guardar_templates:
            # Agregar hoja "Outliers" con los datos
            wb = load_workbook(output_path)

            # Si por algún motivo se genera una hoja llamada "Outliers" más de una vez:
            if "Outliers" in wb.sheetnames:
                del wb["Outliers"]
            ws_outliers = wb.create_sheet("Outliers")

            ddg_headers = [f"Outlier_∆∆G-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]]
            ee_headers = [f"Outlier_%ee-{x:.2f}%" for x in [i * 0.01 for i in range(1, 51)]]

            ws_outliers.append(["Model"] + ddg_headers)
            ws_outliers.append([model_name] + outliers_ddg)
            ws_outliers.append([])
            ws_outliers.append(["Model"] + ee_headers)
            ws_outliers.append([model_name] + outliers_ee)

            wb.save(output_path)
            wb.close()

            logging.info(f"Template de {model_name} guardado con hoja Outliers.")

# -------------------- FUNCIÓN 3 --------------------
def extract_statistics_from_txt(txt_path):
    stats = {"R2": None, "RMSE": None, "MAE": None, "MAPE": None}
    if not os.path.exists(txt_path):
        return stats

    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()[::-1]
        for line in lines:
            line = line.strip()
            if line.startswith("R2") and stats["R2"] is None:
                match = re.search(r"R2:\s*(-?\d+\.\d+)", line)
                if match:
                    stats["R2"] = float(match.group(1))
            elif "RMSE" in line and stats["RMSE"] is None:
                match = re.search(r"RMSE:\s*(\d+\.\d+)", line)
                if match:
                    stats["RMSE"] = float(match.group(1))
            elif "MAE" in line and stats["MAE"] is None:
                match = re.search(r"MAE:\s*(\d+\.\d+%)", line)
                if match:
                    stats["MAE"] = match.group(1)
            elif "MAPE" in line and stats["MAPE"] is None:
                match = re.search(r"MAPE:\s*(\d+\.\d+%)", line)
                if match:
                    stats["MAPE"] = match.group(1)
            # Si ya se encontraron todos, cortar el bucle
            if all(val is not None for val in stats.values()):
                break

    return stats

# -------------------- FUNCIÓN 4 --------------------
def create_master_table(template_folder, output_path, base_folder, modelos_con_error):
    headers_main = [
        "# Entry", "Modelo", "R2", "MAE", "RMSE", "MAPE"
    ] + [f"Outlier_∆∆G-{x:.1f}" for x in [0.1, 0.2, 0.3, 0.4, 0.5]] + [f"Outlier_%ee-{x:.0f}%" for x in [1, 2, 3, 4, 5]]
    
    # Leer JSON por líneas (JSONL) y combinar en un solo diccionario
    outlier_dict_ddg = {}
    with open(os.path.join(base_folder, "outlier_dict_ddg.txt"), 'r', encoding='utf-8') as f_ddg:
        for line in f_ddg:
            outlier_dict_ddg.update(json.loads(line))

    outlier_dict_ee = {}
    with open(os.path.join(base_folder, "outlier_dict_ee.txt"), 'r', encoding='utf-8') as f_ee:
        for line in f_ee:
            outlier_dict_ee.update(json.loads(line))

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Resumen"
    ws_main.append(headers_main)

    ws_ddg = wb.create_sheet("Outliers_∆∆G")
    ws_ddg.append(["# Entry", "Modelo"] + [f"Outlier_∆∆G-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]])

    ws_ee = wb.create_sheet("Outliers_%ee")
    ws_ee.append(["# Entry", "Modelo"] + [f"Outlier_%ee-{x:.1f}%" for x in [i * 0.01 for i in range(1, 51)]])

    entry_number = 1
    for filename in os.listdir(template_folder):
        if filename.endswith(".xlsx"):
            model_name = filename.replace("Template_", "").replace(".xlsx", "")
            txt_path = os.path.join(base_folder, model_name, "Statistics.txt")
            stats = extract_statistics_from_txt(txt_path)

            row_main = [entry_number, model_name, stats["R2"], stats["MAE"], stats["RMSE"], stats["MAPE"]]
            row_main += outlier_dict_ddg.get(model_name, [""] * 5)
            row_main += outlier_dict_ee.get(model_name, [""] * 5)
            ws_main.append(row_main)

            # Aplicar color si el modelo está en lista de errores
            HIGHLIGHT_COLOR = "FFFF00"   # Amarillo
            if model_name in modelos_con_error:
                cell = ws_main.cell(row=ws_main.max_row, column=2)  # Columna "Modelo"
                cell.fill = PatternFill(start_color=HIGHLIGHT_COLOR, end_color=HIGHLIGHT_COLOR, fill_type="solid")

            # Agregar a otras hojas
            row_ddg = [entry_number, model_name] + outlier_dict_ddg.get(model_name, [""] * 50)
            ws_ddg.append(row_ddg)

            row_ee = [entry_number, model_name] + outlier_dict_ee.get(model_name, [""] * 50)
            ws_ee.append(row_ee)

            entry_number += 1

    wb.save(output_path)
    logging.info(f"Archivo Master_Table generado: {output_path}")

# -------------------- FUNCIÓN PRINCIPAL --------------------
def main():
    start_time = time.time()

    base_folder = input("Ruta de la carpeta con subcarpetas de modelos: ").strip()
    logging.info("Verificando estado de finalización de modelos...")

    # Registro de logging
    lineas_criterio = pedir_lineas_criterio()
    log_folder = input("Ruta de la carpeta con LOGS de los modelos: ").strip()
    modelos_ok, modelos_error = obtener_listas_de_modelos(log_folder, lineas_criterio)

    logging.info(f"{len(modelos_ok)} modelos finalizaron correctamente.")
    logging.info(f"{len(modelos_error)} modelos finalizaron con errores.")

    file_prefix = input("Prefijo de los archivos Excel (ej: results): ").strip()
    template_path = input("Ruta del archivo Template base: ").strip()
    guardar_templates = input("¿Desea guardar los archivos Template por modelo? (s/n): ").strip().lower() == 's'

    target_column = input("Columna donde copiar los datos en el Template (ej: B): ").strip().upper()
    source_column = input("Columna a extraer de los Excels fuente (ej: A): ").strip().upper()

    # Elegir carpeta para outputs
    if guardar_templates:
        output_folder = os.path.join(base_folder, "1_Templates_Generados")
        os.makedirs(output_folder, exist_ok=True)
    else:
        output_folder = os.path.join(base_folder, "0_temp_output")
        os.makedirs(output_folder, exist_ok=True)

    while True:
        fila = input("¿En qué número de fila se encuentran los valores a extraer/modificar?: ").strip()
        try:
            fila = int(fila)
            break
        except ValueError:
            print("Debe ser un número entero válido.")

    confirm = input(f"Se modificarán dinámicamente las celdas L{fila} y W{fila}, y se tomarán los valores de K{fila} y V{fila}. ¿Es correcto? (s/n): ").strip().lower()
    if confirm != 's':
        l_cell = input("Indique celda para modificar Outlier ΔΔG (ej: L54): ").strip().upper()
        w_cell = input("Indique celda para modificar Outlier %ee (ej: W54): ").strip().upper()
        k_cell = input("Indique celda de extracción para Outlier ΔΔG (ej: K54): ").strip().upper()
        v_cell = input("Indique celda de extracción para Outlier %ee (ej: V54): ").strip().upper()
    else:
        l_cell = f"L{fila}"
        w_cell = f"W{fila}"
        k_cell = f"K{fila}"
        v_cell = f"V{fila}"

    # Extraer información de los Excels "results" de las subcarpetas y guardarla en un diccionario
    dict_path = os.path.join(base_folder, "results_models_dict.txt")
    if not os.path.exists(dict_path):
        logging.info("Extrayendo datos de los Excels fuente...")
        column_data = get_column_from_excels_in_subfolders(base_folder, file_prefix, source_column)

        with open(dict_path, 'w', encoding='utf-8') as f:
            for model_name, (values, subfolder) in column_data.items():
                json.dump({model_name: {"values": values, "subfolder": subfolder}}, f, ensure_ascii=False)
                f.write('\n')
        logging.info(f"Diccionario guardado en {dict_path}")
    else:
        logging.info(f"Ya existe {dict_path}, cargando datos desde allí.")
        column_data = {}
        with open(dict_path, 'r', encoding='utf-8') as f:
            for line in f:
                entry = json.loads(line)
                for k, v in entry.items():
                    column_data[k] = (v["values"], v["subfolder"])


    logging.info("Generando Templates con datos y cálculo de outliers...")
    generate_templates_and_calculate_outliers(
        template_path, column_data, target_column, output_folder, l_cell, w_cell, k_cell, v_cell, guardar_templates
    )

    master_output_path = input("Ruta donde guardar el archivo Master_Table.xlsx: ").strip()
    master_output_file = os.path.join(master_output_path, "Master_Table.xlsx")

    # Cargar modelos con error desde Resultado General.txt
    ruta_resultado = os.path.join(base_folder, "Resultado General.txt")
    modelos_con_error = []
    if os.path.exists(ruta_resultado):
        with open(ruta_resultado, 'r', encoding='utf-8') as f:
            en_errores = False
            for linea in f:
                if "NO cumplen" in linea:
                    en_errores = True
                    continue
                if en_errores and linea.strip():
                    modelos_con_error.append(linea.strip().replace(".txt", ""))

    logging.info("Generando archivo Master_Table con estadísticas y outliers...")
    create_master_table(output_folder, master_output_file, base_folder, modelos_con_error)

    # Notificación de finalización del script
    winsound.MessageBeep(winsound.MB_ICONASTERISK)

    # Registrar duración total del script
    end_time = time.time()
    duration = end_time - start_time

    days = int(duration // 86400)
    hours = int((duration % 86400) // 3600)
    minutes = int((duration % 3600) // 60)
    seconds = int(duration % 60)

    mensaje_duracion = "Tiempo total: "
    if days > 0:
        mensaje_duracion += f"{days}d {hours}h {minutes}m {seconds}s"
    elif hours > 0:
        mensaje_duracion += f"{hours}h {minutes}m {seconds}s"
    elif minutes > 0:
        mensaje_duracion += f"{minutes}m {seconds}s"
    else:
        mensaje_duracion += f"{seconds}s"

    logging.info(mensaje_duracion)

if __name__ == "__main__":
    main()
