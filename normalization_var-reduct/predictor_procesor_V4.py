# ----------------------------------------------------------
# Librerías necesarias:
#     - openpyxl
#     - xlcalculator
# ----------------------------------------------------------

import os
import logging
import json
import re
import time
import argparse
import tempfile
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill
from xlcalculator import ModelCompiler, Evaluator

# -------------------- CONFIGURACIÓN DE LOGGING --------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)

# -------------------- CONFIGURACIÓN DE ARGPARSE --------------------
def parse_args():
    parser = argparse.ArgumentParser(description="Procesamiento de la predicción configurado por SBATCH")

    parser.add_argument("--base_folder", required=True, help="Ruta base con subcarpetas de modelos")
    parser.add_argument("--output_folder", required=True, help="Carpeta donde guardar Master_Table.xlsx")
    parser.add_argument("--sheet_name", required=True, help="Nombre de la hoja en la plantilla Excel")
    parser.add_argument("--file_prefix", required=True, help="Prefijo de archivos Excel")
    parser.add_argument("--template_path", required=True, help="Ruta del archivo Template base")
    parser.add_argument("--guardar_templates", action="store_true", help="Flag para guardar archivos Template")
    parser.add_argument("--target_column", required=True, help="Columna target en Template (ej: B)")
    parser.add_argument("--source_column", required=True, help="Columna a extraer de Excels fuente (ej: A)")
    parser.add_argument("--log_folder", required=True, help="Carpeta con logs de modelos")
    parser.add_argument("--lineas_criterio", nargs="+", required=True, help="Lista de líneas de finalización")
    parser.add_argument("--fila_datos", required=True, type=int, help="Fila de datos desde donde extraer en el Excel")

    return parser.parse_args()

# -------------------- UTILIDADES LAST_LINES_SELECTOR.PY --------------------
def obtener_listas_de_modelos(ruta_base, lineas_criterio):
    cumplen, no_cumplen = analizar_archivos_en_carpeta(ruta_base, lineas_criterio)
    modelos_ok = [archivo.replace(".txt", "") for archivo in cumplen]
    modelos_error = [archivo.replace(".txt", "") for archivo in no_cumplen]
    return modelos_ok, modelos_error

def analizar_archivos_en_carpeta(ruta_base, lineas_criterio):
    cumplen = []
    no_cumplen = []

    for archivo in os.listdir(ruta_base):
        if archivo.endswith(".txt") and archivo != "log_analysis.txt":
            ruta_completa = os.path.join(ruta_base, archivo)
            if verificar_lineas_finales(ruta_completa, lineas_criterio):
                cumplen.append(archivo)
            else:
                no_cumplen.append(archivo)

    return cumplen, no_cumplen

def verificar_lineas_finales(ruta_archivo, lineas_objetivo):
    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            contenido = f.readlines()
            contenido = [linea.strip() for linea in contenido if linea.strip()]  # eliminamos líneas vacías
            if len(contenido) < len(lineas_objetivo):
                return False
            return contenido[-len(lineas_objetivo):] == lineas_objetivo
    except Exception as e:
        print(f"Error leyendo {ruta_archivo}: {e}")
        return False


# -------------------- UTILIDAD 1: Preparar carpetas de salida --------------------
def preparar_carpeta_output(base_folder: str, guardar_templates: bool):
    """
    Dependiendo de guardar_templates, crea y devuelve:
      - carpeta_templates: donde se guardarán los Template_{modelo}.xlsx (si aplica)
      - carpeta_outliers_txt: donde se guardarán outlier_dict_ddg.txt y outlier_dict_ee.txt
    """
    if guardar_templates:
        carpeta_templates = os.path.join(base_folder, "1_Templates_Models")
        os.makedirs(carpeta_templates, exist_ok=True)

        carpeta_temp = os.path.join(base_folder, "0_temp_output")
        os.makedirs(carpeta_temp, exist_ok=True)

        return carpeta_templates, carpeta_temp
    else:
        carpeta_temp = os.path.join(base_folder, "0_temp_output")
        os.makedirs(carpeta_temp, exist_ok=True)

        return None, carpeta_temp


# -------------------- UTILIDAD 2: Cargar o crear el diccionario de modelos --------------------
def cargar_o_crear_dict_modelos(
    base_folder: str,
    temp_folder: str,
    file_prefix: str,
    source_column: str,
    get_column_fn
):
    """
    - Si existe 'results_models_dict.txt' en base_folder, lo lee y devuelve column_data.
    - Si no, llama a get_column_fn(...) para generar column_data, lo guarda como JSONL y lo devuelve.
    column_data: dict[str, (list, subfolder)]
    """
    dict_path = os.path.join(temp_folder, "results_models_dict.txt")
    if os.path.exists(dict_path):
        logging.info(f"Ya existe {dict_path}, cargando datos desde allí.")
        column_data = {}
        with open(dict_path, 'r', encoding='utf-8') as f:
            for line in f:
                entry = json.loads(line)
                for subfolder, v in entry.items():
                    column_data[subfolder] = (v["values"], v["subfolder"])
        return column_data
    else:
        logging.info("Extrayendo datos de los Excels fuente...")
        column_data = get_column_fn(base_folder, file_prefix, source_column)
        with open(dict_path, 'w', encoding='utf-8') as f:
            for filename, (values, subfolder) in column_data.items():
                json.dump({ subfolder: {"values": values, "subfolder": subfolder}}, f, ensure_ascii=False)
                f.write('\n')
        logging.info(f"Diccionario guardado en {dict_path}")
        return column_data


# -------------------- UTILIDAD 3: Pedir celdas de Outlier --------------------
def pedir_celdas_outlier(fila: int = None):
    """
    Devuelve (l_cell, w_cell, k_cell, v_cell).
    """
    l_cell = f"L{fila}"
    w_cell = f"W{fila}"
    k_cell = f"K{fila}"
    v_cell = f"V{fila}"

    return l_cell, w_cell, k_cell, v_cell


# -------------------- UTILIDAD 4: Extraer columna de Excel en subcarpetas --------------------
def get_column_from_excels_in_subfolders(base_folder: str, file_prefix: str, column_letter: str):
    """
    Recorre cada subcarpeta en base_folder (ignorando carpetas que empiecen 'error' o 'logs'),
    busca archivos que empiecen con file_prefix y no contengan 'training' en su nombre,
    extrae todos los valores de la columna column_letter (hasta celda vacía),
    y devuelve un diccionario {filename: (values_list, subfolder_name)}.
    """
    column_data = {}
    col_idx = column_index_from_string(column_letter)

    for subfolder in os.listdir(base_folder):
        if subfolder.lower().startswith("error") or subfolder.lower().startswith("logs"):
            continue
        subfolder_path = os.path.join(base_folder, subfolder)
        if not os.path.isdir(subfolder_path):
            continue

        for filename in os.listdir(subfolder_path):
            if (
                filename.startswith(file_prefix)
                and filename.endswith(".xlsx")
                and "training" not in filename.lower() # Clave para no agarrar archivos de entrenamiento, sino de validación
            ):
                path = os.path.join(subfolder_path, filename)
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                values = []
                row = 1
                while True:
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value is None:
                        break
                    values.append(cell_value)
                    row += 1
                wb.close()
                column_data[subfolder] = (values, subfolder)
                logging.info(f"Extraídos {len(values)} valores de '{filename}' en columna '{column_letter}'.")
    return column_data


# -------------------- UTILIDAD 5: Calcular Outliers con xlcalculator (guardando cada copia) --------------------
def calcular_outliers_xlcalculator(
    excel_path: str,
    sheet_name: str,
    l_cell: str,
    w_cell: str,
    k_cell: str,
    v_cell: str,
    pasos: int = 50
):
    """
    Abre la planilla con openpyxl + xlcalculator, 
    itera sobre valores de cutoff,
    inyecta los valores en las celdas indicadas,
    recalcula las fórmulas y guarda los resultados.
    """
    outliers_ddg = []
    outliers_ee = []

    # 1) Compilar la planilla como modelo xlcalculator
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(excel_path)
    evaluator = Evaluator(model)

    # Verificar celdas clave antes de empezar
    print("\nValores iniciales:")
    print(f"L54: {evaluator.evaluate(f'{sheet_name}!{l_cell}')}")
    print(f"W54: {evaluator.evaluate(f'{sheet_name}!{w_cell}')}")
    print(f"K54: {evaluator.evaluate(f'{sheet_name}!{k_cell}')}")
    print(f"V54: {evaluator.evaluate(f'{sheet_name}!{v_cell}')}")
    # Verificar algunas celdas de datos
    print("\nVerificación de datos (primeras 3 filas):")
    for row in range(2, 5):
        print(f"Fila {row}:")
        print(f"  C{row}: {evaluator.evaluate(f'{sheet_name}!C{row}')}")
        print(f"  H{row}: {evaluator.evaluate(f'{sheet_name}!H{row}')}")
        print(f"  S{row}: {evaluator.evaluate(f'{sheet_name}!S{row}')}")

    # 2) Iterar cada paso de cutoff (1..pasos)
    for i in range(pasos):
        out_val = round((i + 1) * 0.01, 2)

        # 2a) Inyectar cutoff en l_cell y w_cell
        # CORRECCIÓN: Usar notación "Hoja!Celda" en el primer parámetro
        evaluator.set_cell_value(f"{sheet_name}!{l_cell}", out_val)
        evaluator.set_cell_value(f"{sheet_name}!{w_cell}", out_val)

        # Forzar recálculo completo
        evaluator.evaluate(f"{sheet_name}!A1:Z100")  # Rango amplio para asegurar recálculo
        
        # Obtener resultados con manejo de errores
        try:
            out_ddg = evaluator.evaluate(f"{sheet_name}!{k_cell}") or 0
            out_ee = evaluator.evaluate(f"{sheet_name}!{v_cell}") or 0
        except:
            out_ddg, out_ee = 0, 0

        # Debug
        print(f"\nCutoff: {out_val:.2f}")
        print(f"  Valores intermedios:")
        print(f"  H2: {evaluator.evaluate(f'{sheet_name}!H2')}")
        print(f"  S2: {evaluator.evaluate(f'{sheet_name}!S2')}")
        print(f"  Resultados - K: {out_ddg}, V: {out_ee}")

        if out_ddg is None or out_ee is None:
            logging.warning(f"Valores nulos para cutoff={out_val}. Revisar fórmula.")

        outliers_ddg.append(float(out_ddg))
        outliers_ee.append(float(out_ee))

    return outliers_ddg, outliers_ee


# -------------------- UTILIDAD 6: Calcular Outliers sin crear un nuevo .xlsx con xlcalculator --------------------
def calcular_outliers_sin_guardar_plantilla_xlcalculator(
    template_path: str,
    sheet_name: str,
    values: list,
    target_col_idx: int,
    l_cell: str,
    w_cell: str,
    k_cell: str,
    v_cell: str,
    pasos: int
):
    """
    Similar a la versión anterior, pero:
    - Inserta los valores de 'values' en la columna target_col_idx usando openpyxl.
    - Guarda temporalmente en disco o lo procesa directamente.
    - Luego carga con xlcalculator y calcula los outliers.
    """
    # 1) Abrir plantilla y escribir los valores en la columna target_col_idx
    wb = load_workbook(template_path)
    ws = wb.active
    for i, val in enumerate(values, start=1):
        ws.cell(row=i, column=target_col_idx).value = val

    # Guardar como archivo temporal
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        temp_path = tmp.name
    wb.save(temp_path)
    wb.close()

    # 2) Compilar como modelo xlcalculator
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(temp_path)
    evaluator = Evaluator(model)

    outliers_ddg = []
    outliers_ee = [] 

    # 3) Iterar cada paso de cutoff
    for i in range(pasos):
        out_val = round((i + 1) * 0.01, 2)

        evaluator.set_cell_value(f"{sheet_name}!{l_cell}", out_val)
        evaluator.set_cell_value(f"{sheet_name}!{w_cell}", out_val)

        # Forzar recálculo completo
        evaluator.evaluate(f"{sheet_name}!A1:Z100")  # Rango amplio para asegurar recálculo
        
        # Obtener resultados con manejo de errores
        try:
            out_ddg = evaluator.evaluate(f"{sheet_name}!{k_cell}") or 0
            out_ee = evaluator.evaluate(f"{sheet_name}!{v_cell}") or 0
        except:
            out_ddg, out_ee = 0, 0

        if out_ddg is None or out_ee is None:
            logging.warning(f"Valores nulos para cutoff={out_val}. Revisar fórmula.")

        outliers_ddg.append(float(out_ddg))
        outliers_ee.append(float(out_ee))

    os.remove(temp_path)

    return outliers_ddg, outliers_ee


# -------------------- UTILIDAD 7: Guardar Outliers en JSONL --------------------
def guardar_outliers_jsonl(
    carpeta_outliers: str,
    model_name: str,
    outliers_ddg: list,
    outliers_ee: list
):
    """
    Escribe una línea JSON por modelo en:
      - carpeta_outliers/outlier_dict_ddg.txt
      - carpeta_outliers/outlier_dict_ee.txt
    """
    ruta_ddg = os.path.join(carpeta_outliers, "outlier_dict_ddg.txt")
    ruta_ee = os.path.join(carpeta_outliers, "outlier_dict_ee.txt")

    # Convertir todos los valores a tipos nativos de Python
    ddg_serializable = [float(x) if hasattr(x, 'value') else x for x in outliers_ddg]
    ee_serializable = [float(x) if hasattr(x, 'value') else x for x in outliers_ee]

    with open(ruta_ddg, 'a', encoding='utf-8') as f_ddg:
        f_ddg.write(json.dumps({model_name: ddg_serializable}, ensure_ascii=False) + '\n')

    with open(ruta_ee, 'a', encoding='utf-8') as f_ee:
        f_ee.write(json.dumps({model_name: ee_serializable}, ensure_ascii=False) + '\n')


# -------------------- UTILIDAD 8: Guardar plantilla con hoja “Outliers” (modo guardar plantillas) --------------------
def guardar_plantilla_con_outliers(
    carpeta_templates: str,
    model_name: str,
    outliers_ddg: list,
    outliers_ee: list
):
    """
    Dado que ya existe 'carpeta_templates/Template_{model_name}.xlsx', le agrega (o reemplaza)
    la hoja 'Outliers' con dos bloques de datos:
      - Filas de Outlier_∆∆G-0.01 .. 0.50
      - Filas de Outlier_%ee-0.01 .. 0.50
    """
    ruta_plantilla = os.path.join(carpeta_templates, f"Template_{model_name}.xlsx")
    wb = load_workbook(ruta_plantilla)

    # Eliminar hoja existente si existe
    if "Outliers" in wb.sheetnames:
        del wb["Outliers"]

    ws_outliers = wb.create_sheet("Outliers")

    # Preparar datos
    ddg_headers = [f"Outlier_∆∆G-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]]
    ee_headers = [f"Outlier_%ee-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]]

    ws_outliers.append(["Model"] + ddg_headers)
    ws_outliers.append([model_name] + [float(x) if hasattr(x, 'value') else x for x in outliers_ddg])
    ws_outliers.append([])
    ws_outliers.append(["Model"] + ee_headers)
    ws_outliers.append([model_name] + [float(x) if hasattr(x, 'value') else x for x in outliers_ee])

    wb.save(ruta_plantilla)
    wb.close()


# -------------------- UTILIDAD 9: Extraer estadísticas desde 'Statistics.txt' --------------------
def extract_statistics_from_txt(txt_path: str):
    """
    Extrae métricas específicas de un archivo de texto:
    - R2_LOOCV: bajo 'LOOCV (Evaluación Global)'
    - R2_5FoldCV: bajo '5-Fold CV (Evaluación Global)'
    - R2_ObsVsPred: bajo 'Rendimiento del Conjunto de Validación' -> 'R2 (Observado vs. Predicho)'
    - MAE, RMSE, MAPE: del bloque 'Rendimiento del Conjunto de Validación'
    
    Devuelve:
        dict con las métricas, o None si no se encuentran.
    """
    stats = {
        "R2_LOOCV": None,
        "R2_5FoldCV": None,
        "R2_ObsVsPred": None,
        "MAE": None,
        "RMSE": None,
        "MAPE": None
    }

    if not os.path.exists(txt_path):
        return stats

    current_section = None

    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()

            # Cambiar sección según encabezado detectado
            section_match = re.match(r"^\s*(LOOCV|5-Fold CV|Rendimiento del Conjunto de Validación)", line)
            if section_match:
                section_name = section_match.group(1).strip()
                if "Rendimiento del Conjunto de Validación" in section_name:
                    current_section = "Validation"
                elif "5-Fold CV" in section_name:
                    current_section = "5FoldCV"
                elif "LOOCV" in section_name:
                    current_section = "LOOCV"
            # Analizar líneas dentro de cada sección
            if current_section and current_section == "LOOCV" and stats["R2_LOOCV"] is None:
                match = re.search(r"R2:\s*(-?\d+\.\d+)", line)
                if match:
                    stats["R2_LOOCV"] = float(match.group(1))

            elif current_section and current_section == "5FoldCV" and stats["R2_5FoldCV"] is None:
                match = re.search(r"R2:\s*(-?\d+\.\d+)", line)
                if match:
                    stats["R2_5FoldCV"] = float(match.group(1))

            elif current_section and current_section == "Validation":
                if "R2 (Observado vs. Predicho)" in line and stats["R2_ObsVsPred"] is None:
                    match = re.search(r"R2 \(Observado vs\. Predicho\):\s*(-?\d+\.\d+)", line)
                    if match:
                        stats["R2_ObsVsPred"] = float(match.group(1))
                elif "MAE" in line and stats["MAE"] is None:
                    match = re.search(r"MAE:\s*(\d+\.\d+)", line)
                    if match:
                        stats["MAE"] = float(match.group(1))
                elif "RMSE" in line and stats["RMSE"] is None:
                    match = re.search(r"RMSE:\s*(\d+\.\d+)", line)
                    if match:
                        stats["RMSE"] = float(match.group(1))
                elif "MAPE" in line and stats["MAPE"] is None:
                    match = re.search(r"MAPE:\s*(\d+\.\d+%)", line)
                    if match:
                        stats["MAPE"] = match.group(1)

            # Terminar si se encontró todo
            if all(val is not None for val in stats.values()):
                break

    return stats


# -------------------- UTILIDAD 10: Crear la tabla maestra final --------------------
def create_master_table(
    carpeta_templates: str,
    output_path: str,
    carpeta_outliers_txt: str,
    base_folder: str,
    modelos_con_error: list,
    keys_modelos: list
):
    """
    Genera Master_Table.xlsx con tres hojas:
      - "Resumen" con columnas: #Entry, Modelo, R2_LOOCV, R2_5FoldCV, R2_ObsVsPred, MAE, RMSE, MAPE,
         Outlier_∆∆G-0.1 ... 0.5, Outlier_%ee-1% ... 5%.
      - "Outliers_∆∆G" con columnas: #Entry, Modelo, Outlier_∆∆G-0.01 ... 0.50.
      - "Outliers_%ee" con columnas: #Entry, Modelo, Outlier_%ee-1% ... 5.0%.

    Usa JSONL de outliers (outlier_dict_ddg.txt y outlier_dict_ee.txt)
    y los archivos Template_{model}.xlsx (si existen) para extraer nombres.
    Si carpeta_templates es None o no existe, itera sobre keys_modelos directamente.
    """
    # 1) Leer JSONL de outliers
    outlier_dict_ddg = {}
    ruta_ddg = os.path.join(carpeta_outliers_txt, "outlier_dict_ddg.txt")
    if os.path.exists(ruta_ddg):
        with open(ruta_ddg, 'r', encoding='utf-8') as f_ddg:
            for line in f_ddg:
                outlier_dict_ddg.update(json.loads(line))

    outlier_dict_ee = {}
    ruta_ee = os.path.join(carpeta_outliers_txt, "outlier_dict_ee.txt")
    if os.path.exists(ruta_ee):
        with open(ruta_ee, 'r', encoding='utf-8') as f_ee:
            for line in f_ee:
                outlier_dict_ee.update(json.loads(line))

    # 2) Iniciar libro y hojas
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Resumen"

    headers_main = [
        "# Entry", "Modelo", "R2_LOOCV", "R2_5FoldCV", "R2_ObsVsPred", "MAE", "RMSE", "MAPE"
    ] + [f"Outlier_∆∆G-{x:.1f}" for x in [0.1, 0.2, 0.3, 0.4, 0.5]] \
      + [f"Outlier_%ee-{int(x)}%" for x in [10, 20, 30, 40, 50]]
    ws_main.append(headers_main)

    ws_ddg = wb.create_sheet("Outliers_∆∆G")
    ws_ddg.append(
        ["# Entry", "Modelo"] +
        [f"Outlier_∆∆G-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]]
    )

    ws_ee = wb.create_sheet("Outliers_%ee")
    ws_ee.append(
        ["# Entry", "Modelo"] +
        [f"Outlier_%ee-{x}" for x in [i for i in range(1, 51)]]
    )

    # 3) Determinar lista de modelos
    if carpeta_templates and os.path.isdir(carpeta_templates):
        lista_plantillas = [
            fn for fn in os.listdir(carpeta_templates) if fn.startswith("Template_") and fn.endswith(".xlsx")
        ]
        modelos_list = [fn.replace("Template_", "").replace(".xlsx", "") for fn in lista_plantillas]
    else:
        modelos_list = keys_modelos

    # 4) Poblar hojas
    entry_number = 1
    for model_name in modelos_list:
        # 4.1) Estadísticas desde Statistics.txt
        txt_path = os.path.join(base_folder, model_name, "Statistics.txt")
        stats = extract_statistics_from_txt(txt_path)

        # 4.2) Fila principal
        row_main = [
            entry_number, model_name,
            stats["R2_LOOCV"], stats["R2_5FoldCV"], stats["R2_ObsVsPred"], stats["MAE"], stats["RMSE"], stats["MAPE"]
        ]

        ws_main.append(row_main)

        # 4.3) Resaltar si está en error
        if model_name in modelos_con_error:
            cell = ws_main.cell(row=ws_main.max_row, column=2)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 4.4) Filas de outliers completos
        ddg_vals = outlier_dict_ddg.get(model_name, [""] * 50)
        ee_vals = outlier_dict_ee.get(model_name, [""] * 50)

        ws_ddg.append([entry_number, model_name] + ddg_vals)
        ws_ee.append([entry_number, model_name] + ee_vals)

        entry_number += 1

    # 5.1) Ajustar ancho automático de columnas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Número de columna
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    # 5.2) Agregar en la hoja principal, las referencias alternadas de los outliers a traves de las otras hojas
    agregar_referencias_alternadas(
        wb,
        hoja_principal="Resumen",
        hoja_fuente="Outliers_∆∆G",
        col_destino_inicio=9,  # Columna I
        col_fuente_inicio=12,   # Columna L
        step=10,
        repeticiones=5
    )
    agregar_referencias_alternadas(
        wb,
        hoja_principal="Resumen",
        hoja_fuente="Outliers_%ee",
        col_destino_inicio=14,  # Columna N
        col_fuente_inicio=12,   # Columna L
        step=10,
        repeticiones=5
    )

    # 5.3) Guardar Master_Table.xlsx
    wb.save(output_path)
    logging.info(f"Archivo Master_Table generado: {output_path}")

# -------------------- UTILIDAD 11: Referenciar en excel --------------------
def agregar_referencias_alternadas(
    wb,
    hoja_principal: str,
    hoja_fuente: str,
    col_destino_inicio: int,  # por ejemplo, columna N es 14
    col_fuente_inicio: int,   # por ejemplo, columna R es 18
    step: int = 10,
    repeticiones: int = 5
):
    ws_destino = wb[hoja_principal]
    ws_fuente = wb[hoja_fuente]

    # Determinar última fila ocupada en hoja fuente
    max_row = ws_fuente.max_row

    for i in range(repeticiones):
        col_dest = col_destino_inicio + i * 1
        col_src = col_fuente_inicio + i * step

        col_dest_letter = get_column_letter(col_dest)
        col_src_letter = get_column_letter(col_src)

        for row in range(2, max_row + 1):
            formula = f"='{hoja_fuente}'!{col_src_letter}{row}"
            ws_destino[f"{col_dest_letter}{row}"] = formula


# -------------------- FUNCIÓN PRINCIPAL --------------------
def main():
    start_time = time.time()

    # 1) Inputs básicos
    args = parse_args()

    base_folder = args.base_folder
    master_output_path = args.output_folder
    sheet_name = args.sheet_name
    file_prefix = args.file_prefix
    template_path = args.template_path
    guardar_templates = args.guardar_templates
    target_column = args.target_column
    source_column = args.source_column
    log_folder = args.log_folder
    lineas_criterio = args.lineas_criterio
    fila_datos = args.fila_datos

    logging.info(f"Carpeta base: {base_folder}")
    logging.info(f"Ruta donde guardar el archivo Master_Table.xlsx: {master_output_path}")
    logging.info(f"Nombre de la hoja donde se guardarán los resultados (primera hoja del template.xlsx): {sheet_name}")
    logging.info(f"Prefijo de los archivos Excel (ej: results): {file_prefix}")
    logging.info(f"Ruta del archivo Template base (new_template.xlsx): {template_path}")
    logging.info(f"¿Desea guardar los archivos Template por modelo? (s/n): {guardar_templates}")
    logging.info(f"Columna donde copiar los datos en el Template (ej: B): {target_column}")
    logging.info(f"Columna a extraer de los Excels fuente (ej: A): {source_column}")
    logging.info(f"Carpeta de los archivos log: {log_folder}")
    logging.info(f"Usando líneas de criterio: {lineas_criterio}")
    logging.info(f"Fila de datos: {fila_datos}")

    logging.info("Verificando estado de finalización de modelos...")
    modelos_ok, modelos_error = obtener_listas_de_modelos(log_folder, lineas_criterio)

    logging.info(f"{len(modelos_ok)} modelos finalizaron correctamente.")
    logging.info(f"{len(modelos_error)} modelos finalizaron con errores.")

    # 1.1) Manejo de errores
    if not os.path.exists(base_folder):
        logging.error(f"La carpeta base no existe: {base_folder}")
        exit(1)

    if not os.path.isfile(template_path):
        logging.error(f"El archivo de plantilla no existe: {template_path}")
        exit(1)

    if not os.path.isdir(log_folder):
        logging.error(f"La carpeta de logs no existe: {log_folder}")
        exit(1)

    # 2) Preparar carpetas de salida
    carpeta_templates, carpeta_outliers_txt = preparar_carpeta_output(base_folder, guardar_templates)

    # 3) Pedir celdas de outliers
    l_cell, w_cell, k_cell, v_cell = pedir_celdas_outlier(fila_datos)

    # 4) Cargar o crear el diccionario de modelos (values + subfolder)
    column_data = cargar_o_crear_dict_modelos(
        base_folder,
        carpeta_outliers_txt, # Utilizo la carpeta de outliers para almacenar el diccionario
        file_prefix,
        source_column,
        get_column_from_excels_in_subfolders
    )

    # 5) Preparar JSONL para outliers
    ruta_ddg = os.path.join(carpeta_outliers_txt, "outlier_dict_ddg.txt")
    ruta_ee = os.path.join(carpeta_outliers_txt, "outlier_dict_ee.txt")

    # 6) Iterar modelos y calcular outliers
    try:
        target_col_idx = column_index_from_string(target_column)
    except ValueError:
        logging.error(f"Columna target inválida: {target_column}")
        exit(1)
    pasos = 50

    open(ruta_ddg, 'w').close()
    open(ruta_ee, 'w').close()

    logging.info("Generando Templates (si aplica) y calculando outliers...")
    for model_name, (values, subfolder) in column_data.items():
        if guardar_templates:
            # 6.1a) Copiar template en disco con openpyxl
            wb = load_workbook(template_path)
            ws = wb.active
            for i, val in enumerate(values, start=1):
                ws.cell(row=i, column=target_col_idx, value=val)
            output_path_individual = os.path.join(carpeta_templates, f"Template_{model_name}.xlsx")
            wb.save(output_path_individual)
            wb.close()

            # 6.2a) Calcular outliers sobre esa copia
            outliers_ddg, outliers_ee = calcular_outliers_xlcalculator(
                output_path_individual, sheet_name, l_cell, w_cell, k_cell, v_cell, pasos
            )

            # 6.3a) Guardar JSONL
            guardar_outliers_jsonl(carpeta_outliers_txt, model_name, outliers_ddg, outliers_ee)

            # 6.4a) Añadir hoja "Outliers" a la copia recién creada
            guardar_plantilla_con_outliers(carpeta_templates, model_name, outliers_ddg, outliers_ee)

            logging.info(f"Template de '{model_name}' guardado con hoja Outliers.")

        else:
            # 6.1b) Usar plantilla original con xlwings SIN GUARDAR archivo
            outliers_ddg, outliers_ee = calcular_outliers_sin_guardar_plantilla_xlcalculator(
                template_path, sheet_name, values, target_col_idx, l_cell, w_cell, k_cell, v_cell, pasos
            )
            # 6.2b) Guardar JSONL
            guardar_outliers_jsonl(carpeta_outliers_txt, model_name, outliers_ddg, outliers_ee)

            logging.info(f"Calculados outliers para '{model_name}' (sin guardar plantilla).")


    # 7) Crear Master_Table.xlsx
    os.makedirs(master_output_path, exist_ok=True)
    output_path = os.path.join(master_output_path, "Master_Table.xlsx")

    # 7.1) Llamar a create_master_table
    keys_modelos = list(column_data.keys())
    create_master_table(
        carpeta_templates,
        output_path,
        carpeta_outliers_txt,
        base_folder,
        modelos_error,
        keys_modelos
    )

    # 9) Tiempo total
    end_time = time.time()
    duration = end_time - start_time
    days = int(duration // 86400)
    hours = int((duration % 86400) // 3600)
    minutes = int((duration % 3600) // 60)
    seconds = int(duration % 60)

    mensaje_duracion = "Tiempo total: "
    if days > 0:
        mensaje_duracion += f"{days}d {hours}h {minutes}m {seconds}s"
    elif hours > 0:
        mensaje_duracion += f"{hours}h {minutes}m {seconds}s"
    elif minutes > 0:
        mensaje_duracion += f"{minutes}m {seconds}s"
    else:
        mensaje_duracion += f"{seconds}s"
    logging.info(mensaje_duracion)

    sys.exit(0)


if __name__ == "__main__":
    main()
