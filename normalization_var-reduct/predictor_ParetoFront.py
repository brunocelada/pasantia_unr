import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

"""
----------------------------------------------------------
Frontera de Pareto para modelos combinados
----------------------------------------------------------

Identifica los modelos NO dominados (eficientes) considerando múltiples métricas.
"""

# ----------------------------------------
# Función: ajustar_ancho_columnas_excel
# ----------------------------------------
def ajustar_ancho_columnas_excel(path_excel):
    """
    Ajusta el ancho de las columnas automáticamente para cada hoja del archivo Excel.
    """
    wb = load_workbook(path_excel, data_only=True)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(path_excel)
    logger.info(f"Ajuste de anchos de columnas aplicado: {path_excel}")

# ----------------------------------------
# Configuración básica del logging
# ----------------------------------------
def setup_logging():
    """
    Configura el logging para registrar mensajes en consola con nivel INFO.
    """
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ----------------------------------------
# Función genérica: prompt_input
# Descripción: Solicita al usuario un valor con validación y valor por defecto.
# ----------------------------------------
def prompt_input(message, default=None, validator=None, error_msg="Entrada inválida"):
    """
    Pide al usuario un input con un mensaje personalizado.
    - message: texto que se muestra al usuario.
    - default: valor por defecto si el usuario no ingresa nada.
    - validator: función que recibe el input y retorna True si es válido.
    - error_msg: mensaje de error al reintentar.
    """
    while True:
        try:
            prompt_msg = message
            if default is not None:
                prompt_msg += f" (por defecto '{default}')"
            prompt_msg += ": "
            user_input = input(prompt_msg).strip()
            # Aplica valor por defecto
            result = user_input or default
            # Valida si se proporcionó validador
            if validator and not validator(result):
                raise ValueError(error_msg)
            logger.info(f"Usuario ingresó: {result}")
            return result
        except Exception as e:
            logger.warning(f"{e}. Por favor, intente de nuevo.")

# ----------------------------------------
# Función: prompt_excel_path (usa prompt_input)
# Descripción: Solicita al usuario la ruta del archivo Excel de manera genérica.
# ----------------------------------------
def prompt_excel_path():
    """
    Pide al usuario la ruta del archivo Excel.
    Utiliza prompt_input para manejar validación y valor por defecto.
    """
    return prompt_input(
        message="Ruta del archivo Excel",
        default='C:\\Linux\\Master_Table.xlsx',
        validator=lambda p: bool(p and os.path.splitext(p)[1] in ['.xlsx', '.xls']),
        error_msg="Debe ingresar un nombre de archivo válido con extensión .xlsx o .xls"
    )

def load_master_table(filepath: str, sheet_name: str = None) -> pd.DataFrame:
    """
    Carga el archivo Master_Table.
    """
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    return df

def load_master_table(filepath: str, sheet_name: str = None) -> pd.DataFrame:
    """
    Carga el archivo Master_Table.
    """
    return pd.read_excel(filepath, sheet_name=sheet_name)

def parse_model_column(df: pd.DataFrame, model_col: str = "Modelo") -> pd.DataFrame:
    """
    Divide la columna 'Modelo' en Normalization, Reduction y Model.
    """
    split_cols = df[model_col].str.split('_', expand=True)
    df['Normalization'] = split_cols[0]
    df['Reduction'] = split_cols[1]
    df['Model'] = split_cols[2]
    return df

def clean_percentage_columns(df: pd.DataFrame, percentage_columns: list) -> pd.DataFrame:
    """
    Quita '%' y convierte a float.
    """
    for col in percentage_columns:
        if df[col].dtype == object:
            df[col] = df[col].str.rstrip('%').astype(float)
    return df

def is_r2(metric: str) -> bool:
    """
    Detecta si la métrica es un R².
    """
    return "R2" in metric

def is_dominated(point, others, metrics):
    """
    Verifica si un punto está dominado por al menos otro.
    Dominado = Peor o igual en todas las métricas y peor en al menos una.
    """
    for _, other in others.iterrows():
        better_in_all = True
        better_in_at_least_one = False

        for m in metrics:
            if is_r2(m):  # R2: mayor es mejor
                if other[m] < point[m]:
                    better_in_all = False
                    break
                if other[m] > point[m]:
                    better_in_at_least_one = True
            else:  # Errores y outliers: menor es mejor
                if other[m] > point[m]:
                    better_in_all = False
                    break
                if other[m] < point[m]:
                    better_in_at_least_one = True

        if better_in_all and better_in_at_least_one:
            return True

    return False

def compute_pareto_front(df: pd.DataFrame, metrics: list) -> pd.DataFrame:
    """
    Devuelve la frontera de Pareto (no dominados).
    """
    pareto_mask = []
    for idx, row in df.iterrows():
        others = df.drop(idx)
        dominated = is_dominated(row, others, metrics)
        pareto_mask.append(not dominated)

    pareto_df = df[pareto_mask].reset_index(drop=True)
    return pareto_df

def main():
    filepath = prompt_excel_path()
    sheet_name = "Resumen"
    model_col = "Modelo"

    metrics = [
        "R2_LOOCV",
        "R2_5FoldCV",
        "R2_ObsVsPred",
        "MAE",
        "RMSE",
        "MAPE",
        "Outlier_∆∆G-0.3",
        "Outlier_%ee-30%"
    ]

    percentage_columns = ["MAPE"]

    df = load_master_table(filepath, sheet_name)
    df = parse_model_column(df, model_col)
    df = clean_percentage_columns(df, percentage_columns)

    pareto_df = compute_pareto_front(df, metrics)

    print("=== Modelos en la Frontera de Pareto ===")
    print(pareto_df[[model_col] + metrics])

    # Guardar
    output_file = filepath.replace(".xlsx", "_Pareto_Front.xlsx")
    pareto_df.to_excel(output_file, index=False)
    print(f"\n✅ Frontera de Pareto guardada en: {output_file}")

    # Ajusta automáticamente el ancho de las columnas del Excel generado.
    ajustar_ancho_columnas_excel(output_file)

if __name__ == "__main__":
    main()
