# -----------------------------------------------------------------------------------------------------------------------
# Se debe usar este script por el momento en WINDOWS solamente, y tener instalado EXCEL y las librerías necesarias.
# Durante su ejecución, no abrir ni cerrar Excel, ya que podrían ocurrir errores en el script.

# Librerías necesarias:
#     - openpyxl
#     - xlwings

# ------------------------------------------------------------------------------------------------------------------------

import os
import logging
import json
import re
import time
import winsound

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill
import xlwings as xw

from last_lines_selector import obtener_listas_de_modelos, pedir_lineas_criterio


# -------------------- CONFIGURACIÓN DE LOGGING --------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)


# -------------------- UTILIDAD 1: Preparar carpetas de salida --------------------
def preparar_carpeta_output(base_folder: str, guardar_templates: bool):
    """
    Dependiendo de guardar_templates, crea y devuelve:
      - carpeta_templates: donde se guardarán los Template_{modelo}.xlsx (si aplica)
      - carpeta_outliers_txt: donde se guardarán outlier_dict_ddg.txt y outlier_dict_ee.txt
    """
    if guardar_templates:
        carpeta_templates = os.path.join(base_folder, "1_Templates_Models")
        os.makedirs(carpeta_templates, exist_ok=True)

        carpeta_temp = os.path.join(base_folder, "0_temp_output")
        os.makedirs(carpeta_temp, exist_ok=True)

        return carpeta_templates, carpeta_temp
    else:
        carpeta_temp = os.path.join(base_folder, "0_temp_output")
        os.makedirs(carpeta_temp, exist_ok=True)

        return None, carpeta_temp


# -------------------- UTILIDAD 2: Cargar o crear el diccionario de modelos --------------------
def cargar_o_crear_dict_modelos(
    base_folder: str,
    temp_folder: str,
    file_prefix: str,
    source_column: str,
    get_column_fn
):
    """
    - Si existe 'results_models_dict.txt' en base_folder, lo lee y devuelve column_data.
    - Si no, llama a get_column_fn(...) para generar column_data, lo guarda como JSONL y lo devuelve.
    column_data: dict[str, (list, subfolder)]
    """
    dict_path = os.path.join(temp_folder, "results_models_dict.txt")
    if os.path.exists(dict_path):
        logging.info(f"Ya existe {dict_path}, cargando datos desde allí.")
        column_data = {}
        with open(dict_path, 'r', encoding='utf-8') as f:
            for line in f:
                entry = json.loads(line)
                for subfolder, v in entry.items():
                    column_data[subfolder] = (v["values"], v["subfolder"])
        return column_data
    else:
        logging.info("Extrayendo datos de los Excels fuente...")
        column_data = get_column_fn(base_folder, file_prefix, source_column)
        with open(dict_path, 'w', encoding='utf-8') as f:
            for filename, (values, subfolder) in column_data.items():
                json.dump({ subfolder: {"values": values, "subfolder": subfolder}}, f, ensure_ascii=False)
                f.write('\n')
        logging.info(f"Diccionario guardado en {dict_path}")
        return column_data


# -------------------- UTILIDAD 3: Pedir celdas de Outlier --------------------
def pedir_celdas_outlier():
    """
    Pide al usuario la fila base y verifica si usar L{fila}, W{fila}, K{fila}, V{fila}.
    Si no, solicita manualmente cada celda.
    Devuelve (l_cell, w_cell, k_cell, v_cell).
    """
    while True:
        s = input("¿En qué número de fila se encuentran los valores a extraer/modificar?: ").strip()
        try:
            fila = int(s)
            break
        except ValueError:
            print("Debe ser un número entero válido.")

    confirm = input(
        f"Se modificarán dinámicamente las celdas L{fila} y W{fila}, "
        f"y se tomarán los valores de K{fila} y V{fila}. ¿Es correcto? (s/n): "
    ).strip().lower()
    if confirm != 's':
        l_cell = input("Indique celda para modificar Outlier ΔΔG (ej: L54): ").strip().upper()
        w_cell = input("Indique celda para modificar Outlier %ee (ej: W54): ").strip().upper()
        k_cell = input("Indique celda de extracción para Outlier ΔΔG (ej: K54): ").strip().upper()
        v_cell = input("Indique celda de extracción para Outlier %ee (ej: V54): ").strip().upper()
    else:
        l_cell = f"L{fila}"
        w_cell = f"W{fila}"
        k_cell = f"K{fila}"
        v_cell = f"V{fila}"

    return l_cell, w_cell, k_cell, v_cell


# -------------------- UTILIDAD 4: Extraer columna de Excel en subcarpetas --------------------
def get_column_from_excels_in_subfolders(base_folder: str, file_prefix: str, column_letter: str):
    """
    Recorre cada subcarpeta en base_folder (ignorando carpetas que empiecen 'error' o 'logs'),
    busca archivos que empiecen con file_prefix y no contengan 'training' en su nombre,
    extrae todos los valores de la columna column_letter (hasta celda vacía),
    y devuelve un diccionario {filename: (values_list, subfolder_name)}.
    """
    column_data = {}
    col_idx = column_index_from_string(column_letter)

    for subfolder in os.listdir(base_folder):
        if subfolder.lower().startswith("error") or subfolder.lower().startswith("logs"):
            continue
        subfolder_path = os.path.join(base_folder, subfolder)
        if not os.path.isdir(subfolder_path):
            continue

        for filename in os.listdir(subfolder_path):
            if (
                filename.startswith(file_prefix)
                and filename.endswith(".xlsx")
                and "training" not in filename.lower() # Clave para no agarrar archivos de entrenamiento, sino de validación
            ):
                path = os.path.join(subfolder_path, filename)
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                values = []
                row = 1
                while True:
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value is None:
                        break
                    values.append(cell_value)
                    row += 1
                wb.close()
                column_data[subfolder] = (values, subfolder)
                logging.info(f"Extraídos {len(values)} valores de '{filename}' en columna '{column_letter}'.")
    return column_data


# -------------------- UTILIDAD 5: Calcular Outliers con xlwings (guardando cada copia) --------------------
def calcular_outliers_xlwings(
    excel_path: str,
    l_cell: str,
    w_cell: str,
    k_cell: str,
    v_cell: str,
    pasos: int = 50,
    sleep_time: float = 0.2
):
    """
    Abre con xlwings el archivo 'excel_path', para cada iteración 1..pasos:
      - Calcula cutoff = (i * 0.01)
      - Asigna cutoff en l_cell y w_cell
      - Ejecuta book.app.calculate()
      - Espera sleep_time segundos
      - Lee k_cell y v_cell, los agrega a las listas de outliers
    Devuelve dos listas: (outliers_ddg, outliers_ee)
    """
    outliers_ddg = []
    outliers_ee = []
    app = None
    book = None

    try:
        app = xw.App(visible=False)
        book = app.books.open(excel_path)
        sheet = book.sheets[0]

        for i in range(pasos):
            out_val = round((i + 1) * 0.01, 2)
            print(f"Calculando los outliers con un cutoff de: {out_val:.2f}", end='\r')

            sheet.range(l_cell).value = out_val
            sheet.range(w_cell).value = out_val
            book.app.calculate()
            time.sleep(sleep_time)

            outliers_ddg.append(sheet.range(k_cell).value)
            outliers_ee.append(sheet.range(v_cell).value)

    except Exception as e:
        logging.error(f"Error al procesar {excel_path} con xlwings: {e}", exc_info=True)
        return [], []

    finally:
        if book:
            try:
                book.close()
            except Exception as e:
                logging.warning(f"No se pudo cerrar el libro: {e}")
        if app:
            try:
                app.quit()
            except Exception as e:
                logging.warning(f"No se pudo cerrar la app de Excel: {e}")

    return outliers_ddg, outliers_ee


# -------------------- UTILIDAD 6: Calcular Outliers sin crear un nuevo .xlsx --------------------
def calcular_outliers_sin_guardar_plantilla(
    template_path: str,
    values: list,
    target_col_idx: int,
    l_cell: str,
    w_cell: str,
    k_cell: str,
    v_cell: str,
    pasos: int,
    sleep_time: float
):
    """
    Abre directamente la plantilla (template_path) con xlwings, sin guardarla nunca.
    1) Inserta 'values' en la columna target_col_idx.
    2) Itera de 1..pasos: asigna cutoff en l_cell y w_cell, calcula, lee k_cell, v_cell.
    3) Al final cierra sin guardar, dejando intacto template_path en disco.
    Devuelve (outliers_ddg, outliers_ee).
    """
    outliers_ddg = []
    outliers_ee = []
    app = None
    book = None

    try:
        app = xw.App(visible=False)
        book = app.books.open(template_path)
        sheet = book.sheets[0]

        # Insertar los valores en la columna target_col_idx
        col_letra = sheet.api.Columns(target_col_idx).Address.replace("$", "").split(":")[0]
        start_cell = f"{col_letra}1"
        sheet.range(start_cell).options(transpose=True).value = values

        for i in range(pasos):
            out_val = round((i + 1) * 0.01, 2)
            print(f"Calculando los outliers con un cutoff de: {out_val:.2f}", end='\r')

            sheet.range(l_cell).value = out_val
            sheet.range(w_cell).value = out_val
            book.app.calculate()
            time.sleep(sleep_time)

            outliers_ddg.append(sheet.range(k_cell).value)
            outliers_ee.append(sheet.range(v_cell).value)

    except Exception as e:
        logging.error(f"Error al procesar {template_path} con xlwings (sin guardar): {e}", exc_info=True)
        return [], []

    finally:
        if book:
            try:
                book.close()
            except Exception as e:
                logging.warning(f"No se pudo cerrar el libro sin guardar: {e}")
        if app:
            try:
                app.quit()
            except Exception as e:
                logging.warning(f"No se pudo cerrar la app de Excel: {e}")

    return outliers_ddg, outliers_ee


# -------------------- UTILIDAD 7: Guardar Outliers en JSONL --------------------
def guardar_outliers_jsonl(
    carpeta_outliers: str,
    model_name: str,
    outliers_ddg: list,
    outliers_ee: list
):
    """
    Escribe una línea JSON por modelo en:
      - carpeta_outliers/outlier_dict_ddg.txt
      - carpeta_outliers/outlier_dict_ee.txt
    """
    ruta_ddg = os.path.join(carpeta_outliers, "outlier_dict_ddg.txt")
    ruta_ee = os.path.join(carpeta_outliers, "outlier_dict_ee.txt")

    with open(ruta_ddg, 'a', encoding='utf-8') as f_ddg:
        f_ddg.write(json.dumps({model_name: outliers_ddg}, ensure_ascii=False) + '\n')

    with open(ruta_ee, 'a', encoding='utf-8') as f_ee:
        f_ee.write(json.dumps({model_name: outliers_ee}, ensure_ascii=False) + '\n')


# -------------------- UTILIDAD 8: Guardar plantilla con hoja “Outliers” (modo guardar plantillas) --------------------
def guardar_plantilla_con_outliers(
    carpeta_templates: str,
    model_name: str,
    outliers_ddg: list,
    outliers_ee: list
):
    """
    Dado que ya existe 'carpeta_templates/Template_{model_name}.xlsx', le agrega (o reemplaza)
    la hoja 'Outliers' con dos bloques de datos:
      - Filas de Outlier_∆∆G-0.01 .. 0.50
      - Filas de Outlier_%ee-0.01 .. 0.50
    """
    ruta_plantilla = os.path.join(carpeta_templates, f"Template_{model_name}.xlsx")
    wb = load_workbook(ruta_plantilla)
    if "Outliers" in wb.sheetnames:
        del wb["Outliers"]
    ws_outliers = wb.create_sheet("Outliers")

    ddg_headers = [f"Outlier_∆∆G-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]]
    ee_headers = [f"Outlier_%ee-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]]

    ws_outliers.append(["Model"] + ddg_headers)
    ws_outliers.append([model_name] + outliers_ddg)
    ws_outliers.append([])
    ws_outliers.append(["Model"] + ee_headers)
    ws_outliers.append([model_name] + outliers_ee)

    wb.save(ruta_plantilla)
    wb.close()


# -------------------- UTILIDAD 9: Extraer estadísticas desde 'Statistics.txt' --------------------
def extract_statistics_from_txt(txt_path: str):
    """
    Lee las líneas de txt_path (reverso) buscando R2, RMSE, MAE, MAPE.
    Devuelve stats = {"R2": float or None, "RMSE":float or None, "MAE": str or None, "MAPE": str or None}.
    """
    stats = {"R2": None, "RMSE": None, "MAE": None, "MAPE": None}
    if not os.path.exists(txt_path):
        return stats

    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()[::-1]
        for line in lines:
            line = line.strip()
            if line.startswith("R2") and stats["R2"] is None:
                match = re.search(r"R2:\s*(-?\d+\.\d+)", line)
                if match:
                    stats["R2"] = float(match.group(1))
            elif "RMSE" in line and stats["RMSE"] is None:
                match = re.search(r"RMSE:\s*(\d+\.\d+)", line)
                if match:
                    stats["RMSE"] = float(match.group(1))
            elif "MAE" in line and stats["MAE"] is None:
                match = re.search(r"MAE:\s*(\d+\.\d+)", line)
                if match:
                    stats["MAE"] = match.group(1)
            elif "MAPE" in line and stats["MAPE"] is None:
                match = re.search(r"MAPE:\s*(\d+\.\d+%)", line)
                if match:
                    stats["MAPE"] = match.group(1)

            # Si ya se encontraron todos, cortar el bucle
            if all(val is not None for val in stats.values()):
                break
    return stats


# -------------------- UTILIDAD 10: Crear la tabla maestra final --------------------
def create_master_table(
    carpeta_templates: str,
    output_path: str,
    carpeta_outliers_txt: str,
    base_folder: str,
    modelos_con_error: list,
    keys_modelos: list
):
    """
    Genera Master_Table.xlsx con tres hojas:
      - "Resumen" con columnas: #Entry, Modelo, R2, MAE, RMSE, MAPE,
         Outlier_∆∆G-0.1 ... 0.5, Outlier_%ee-1% ... 5%.
      - "Outliers_∆∆G" con columnas: #Entry, Modelo, Outlier_∆∆G-0.01 ... 0.50.
      - "Outliers_%ee" con columnas: #Entry, Modelo, Outlier_%ee-1% ... 5.0%.

    Usa JSONL de outliers (outlier_dict_ddg.txt y outlier_dict_ee.txt)
    y los archivos Template_{model}.xlsx (si existen) para extraer nombres.
    Si carpeta_templates es None o no existe, itera sobre keys_modelos directamente.
    """
    # 1) Leer JSONL de outliers
    outlier_dict_ddg = {}
    ruta_ddg = os.path.join(carpeta_outliers_txt, "outlier_dict_ddg.txt")
    if os.path.exists(ruta_ddg):
        with open(ruta_ddg, 'r', encoding='utf-8') as f_ddg:
            for line in f_ddg:
                outlier_dict_ddg.update(json.loads(line))

    outlier_dict_ee = {}
    ruta_ee = os.path.join(carpeta_outliers_txt, "outlier_dict_ee.txt")
    if os.path.exists(ruta_ee):
        with open(ruta_ee, 'r', encoding='utf-8') as f_ee:
            for line in f_ee:
                outlier_dict_ee.update(json.loads(line))

    # 2) Iniciar libro y hojas
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Resumen"

    headers_main = [
        "# Entry", "Modelo", "R2", "MAE", "RMSE", "MAPE"
    ] + [f"Outlier_∆∆G-{x:.1f}" for x in [0.1, 0.2, 0.3, 0.4, 0.5]] \
      + [f"Outlier_%ee-{int(x)}%" for x in [10, 20, 30, 40, 50]]
    ws_main.append(headers_main)

    ws_ddg = wb.create_sheet("Outliers_∆∆G")
    ws_ddg.append(
        ["# Entry", "Modelo"] +
        [f"Outlier_∆∆G-{x:.2f}" for x in [i * 0.01 for i in range(1, 51)]]
    )

    ws_ee = wb.create_sheet("Outliers_%ee")
    ws_ee.append(
        ["# Entry", "Modelo"] +
        [f"Outlier_%ee-{x}" for x in [i for i in range(1, 51)]]
    )

    # 3) Determinar lista de modelos
    if carpeta_templates and os.path.isdir(carpeta_templates):
        lista_plantillas = [
            fn for fn in os.listdir(carpeta_templates) if fn.startswith("Template_") and fn.endswith(".xlsx")
        ]
        modelos_list = [fn.replace("Template_", "").replace(".xlsx", "") for fn in lista_plantillas]
    else:
        modelos_list = keys_modelos

    # 4) Poblar hojas
    entry_number = 1
    for model_name in modelos_list:
        # 4.1) Estadísticas desde Statistics.txt
        txt_path = os.path.join(base_folder, model_name, "Statistics.txt")
        stats = extract_statistics_from_txt(txt_path)

        # 4.2) Fila principal
        row_main = [
            entry_number, model_name,
            stats["R2"], stats["MAE"], stats["RMSE"], stats["MAPE"]
        ]

        ws_main.append(row_main)

        # 4.3) Resaltar si está en error
        if model_name in modelos_con_error:
            cell = ws_main.cell(row=ws_main.max_row, column=2)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 4.4) Filas de outliers completos
        ddg_vals = outlier_dict_ddg.get(model_name, [""] * 50)
        ee_vals = outlier_dict_ee.get(model_name, [""] * 50)

        ws_ddg.append([entry_number, model_name] + ddg_vals)
        ws_ee.append([entry_number, model_name] + ee_vals)

        entry_number += 1

    # 5.1) Ajustar ancho automático de columnas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Número de columna
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    # 5.2) Agregar en la hoja principal, las referencias alternadas de los outliers a traves de las otras hojas
    agregar_referencias_alternadas(
        wb,
        hoja_principal="Resumen",
        hoja_fuente="Outliers_∆∆G",
        col_destino_inicio=7,  # Columna G
        col_fuente_inicio=12,   # Columna L
        step=10,
        repeticiones=5
    )
    agregar_referencias_alternadas(
        wb,
        hoja_principal="Resumen",
        hoja_fuente="Outliers_%ee",
        col_destino_inicio=12,  # Columna L
        col_fuente_inicio=12,   # Columna L
        step=10,
        repeticiones=5
    )

    # 5.3) Guardar Master_Table.xlsx
    wb.save(output_path)
    logging.info(f"Archivo Master_Table generado: {output_path}")

# -------------------- UTILIDAD 11: Referenciar en excel --------------------
def agregar_referencias_alternadas(
    wb,
    hoja_principal: str,
    hoja_fuente: str,
    col_destino_inicio: int,  # por ejemplo, columna N es 14
    col_fuente_inicio: int,   # por ejemplo, columna R es 18
    step: int = 10,
    repeticiones: int = 5
):
    ws_destino = wb[hoja_principal]
    ws_fuente = wb[hoja_fuente]

    # Determinar última fila ocupada en hoja fuente
    max_row = ws_fuente.max_row

    for i in range(repeticiones):
        col_dest = col_destino_inicio + i * 1
        col_src = col_fuente_inicio + i * step

        col_dest_letter = get_column_letter(col_dest)
        col_src_letter = get_column_letter(col_src)

        for row in range(2, max_row + 1):
            formula = f"='{hoja_fuente}'!{col_src_letter}{row}"
            ws_destino[f"{col_dest_letter}{row}"] = formula


# -------------------- FUNCIÓN PRINCIPAL --------------------
def main():
    start_time = time.time()

    # # 1) Inputs básicos
    base_folder = input("Ruta de la carpeta con subcarpetas de modelos: ").strip()
    logging.info("Verificando estado de finalización de modelos...")

    lineas_criterio = pedir_lineas_criterio()
    log_folder = input("Ruta de la carpeta con LOGS de los modelos: ").strip()
    modelos_ok, modelos_error = obtener_listas_de_modelos(log_folder, lineas_criterio)

    logging.info(f"{len(modelos_ok)} modelos finalizaron correctamente.")
    logging.info(f"{len(modelos_error)} modelos finalizaron con errores.")

    file_prefix = input("Prefijo de los archivos Excel (ej: results): ").strip()
    template_path = input("Ruta del archivo Template base: ").strip()
    guardar_templates = (input("¿Desea guardar los archivos Template por modelo? (s/n): ")
                         .strip().lower() == 's')

    target_column = input("Columna donde copiar los datos en el Template (ej: B): ").strip().upper()
    source_column = input("Columna a extraer de los Excels fuente (ej: A): ").strip().upper()

    # 2) Preparar carpetas de salida
    carpeta_templates, carpeta_outliers_txt = preparar_carpeta_output(base_folder, guardar_templates)

    # 3) Pedir celdas de outliers
    l_cell, w_cell, k_cell, v_cell = pedir_celdas_outlier()

    # 4) Cargar o crear el diccionario de modelos (values + subfolder)
    column_data = cargar_o_crear_dict_modelos(
        base_folder,
        carpeta_outliers_txt, # Utilizo la carpeta de outliers para almacenar el diccionario
        file_prefix,
        source_column,
        get_column_from_excels_in_subfolders
    )

    # 5) Preparar JSONL para outliers
    ruta_ddg = os.path.join(carpeta_outliers_txt, "outlier_dict_ddg.txt")
    ruta_ee = os.path.join(carpeta_outliers_txt, "outlier_dict_ee.txt")
    open(ruta_ddg, 'w').close()
    open(ruta_ee, 'w').close()

    # 6) Iterar modelos y calcular outliers
    target_col_idx = column_index_from_string(target_column)
    pasos = 50
    sleep_time = 0.2

    logging.info("Generando Templates (si aplica) y calculando outliers...")
    for model_name, (values, subfolder) in column_data.items():
        if guardar_templates:
            # 6.1a) Copiar template en disco con openpyxl
            wb = load_workbook(template_path)
            ws = wb.active
            for i, val in enumerate(values, start=1):
                ws.cell(row=i, column=target_col_idx, value=val)
            output_path_individual = os.path.join(carpeta_templates, f"Template_{model_name}.xlsx")
            wb.save(output_path_individual)
            wb.close()

            # 6.2a) Calcular outliers sobre esa copia
            outliers_ddg, outliers_ee = calcular_outliers_xlwings(
                output_path_individual, l_cell, w_cell, k_cell, v_cell, pasos, sleep_time
            )

            # 6.3a) Guardar JSONL
            guardar_outliers_jsonl(carpeta_outliers_txt, model_name, outliers_ddg, outliers_ee)

            # 6.4a) Añadir hoja "Outliers" a la copia recién creada
            guardar_plantilla_con_outliers(carpeta_templates, model_name, outliers_ddg, outliers_ee)

            logging.info(f"Template de '{model_name}' guardado con hoja Outliers.")

        else:
            # 6.1b) Usar plantilla original con xlwings SIN GUARDAR archivo
            outliers_ddg, outliers_ee = calcular_outliers_sin_guardar_plantilla(
                template_path, values, target_col_idx, l_cell, w_cell, k_cell, v_cell, pasos, sleep_time
            )
            # 6.2b) Guardar JSONL
            guardar_outliers_jsonl(carpeta_outliers_txt, model_name, outliers_ddg, outliers_ee)

            logging.info(f"Calculados outliers para '{model_name}' (sin guardar plantilla).")

    # 7) Crear Master_Table.xlsx
    winsound.MessageBeep(winsound.MB_ICONASTERISK)
    master_output_path = input("Ruta donde guardar el archivo Master_Table.xlsx: ").strip()
    os.makedirs(master_output_path, exist_ok=True)
    global output_path
    output_path = os.path.join(master_output_path, "Master_Table.xlsx")

    # 7.1) Llamar a create_master_table
    keys_modelos = list(column_data.keys())
    create_master_table(
        carpeta_templates,
        output_path,
        carpeta_outliers_txt,
        base_folder,
        modelos_error,
        keys_modelos
    )

    # -------------------------------------------------------------------------------------
    # 8) Si no se guardaron plantillas, borrar carpeta temporal + JSONL
    # (HECHA COMENTARIO PARA EVITAR BORRAR LOS DICCIONARIOS DE OUTLIERS en proceso de debug)
    
    # if not guardar_templates:
    #     # Borrar los archivos JSONL
    #     try:
    #         os.remove(ruta_ddg)
    #         os.remove(ruta_ee)
    #         logging.info("Archivos JSONL temporales eliminados.")
    #     except Exception as e:
    #         logging.warning(f"No se pudo eliminar JSONL temporal: {e}")
    #     # Borrar carpeta si quedó vacía
    #     try:
    #         os.rmdir(carpeta_outliers_txt)
    #         logging.info("Carpeta temporal eliminada.")
    #     except Exception as e:
    #         logging.warning(f"No se pudo eliminar la carpeta temporal: {e}")
    # -------------------------------------------------------------------------------------

    # 9) Notificación final y tiempo total
    end_time = time.time()
    duration = end_time - start_time
    days = int(duration // 86400)
    hours = int((duration % 86400) // 3600)
    minutes = int((duration % 3600) // 60)
    seconds = int(duration % 60)

    mensaje_duracion = "Tiempo total: "
    if days > 0:
        mensaje_duracion += f"{days}d {hours}h {minutes}m {seconds}s"
    elif hours > 0:
        mensaje_duracion += f"{hours}h {minutes}m {seconds}s"
    elif minutes > 0:
        mensaje_duracion += f"{minutes}m {seconds}s"
    else:
        mensaje_duracion += f"{seconds}s"
    logging.info(mensaje_duracion)

    winsound.MessageBeep(winsound.MB_ICONASTERISK)

if __name__ == "__main__":
    main()
