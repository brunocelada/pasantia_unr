import os
import json
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ----------------------------------------
# Configuración básica del logging
# ----------------------------------------
def setup_logging():
    """
    Configura el logging para registrar mensajes en consola con nivel INFO.
    """
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ----------------------------------------
# Función genérica: prompt_input
# Descripción: Solicita al usuario un valor con validación y valor por defecto.
# ----------------------------------------
def prompt_input(message, default=None, validator=None, error_msg="Entrada inválida"):
    """
    Pide al usuario un input con un mensaje personalizado.
    - message: texto que se muestra al usuario.
    - default: valor por defecto si el usuario no ingresa nada.
    - validator: función que recibe el input y retorna True si es válido.
    - error_msg: mensaje de error al reintentar.
    """
    while True:
        try:
            prompt_msg = message
            if default is not None:
                prompt_msg += f" (por defecto '{default}')"
            prompt_msg += ": "
            user_input = input(prompt_msg).strip()
            # Aplica valor por defecto
            result = user_input or default
            # Valida si se proporcionó validador
            if validator and not validator(result):
                raise ValueError(error_msg)
            logger.info(f"Usuario ingresó: {result}")
            return result
        except Exception as e:
            logger.warning(f"{e}. Por favor, intente de nuevo.")

# ----------------------------------------
# Función: prompt_excel_path (usa prompt_input)
# Descripción: Solicita al usuario la ruta del archivo Excel de manera genérica.
# ----------------------------------------
def prompt_excel_path():
    """
    Pide al usuario la ruta del archivo Excel.
    Utiliza prompt_input para manejar validación y valor por defecto.
    """
    return prompt_input(
        message="Ruta del archivo Excel",
        default='C:\\Linux\\Master_Table.xlsx',
        validator=lambda p: bool(p and os.path.splitext(p)[1] in ['.xlsx', '.xls']),
        error_msg="Debe ingresar un nombre de archivo válido con extensión .xlsx o .xls"
    )

# ----------------------------------------
# Función: prompt_remove_columns
# Descripción: Permite al usuario eliminar descriptores de la lista antes del análisis.
# ----------------------------------------
def prompt_remove_columns(descriptors):
    """
    Muestra las columnas disponibles y permite al usuario eliminar las que no desea analizar.
    Retorna la lista de descriptores restantes.
    """
    while True:
        logger.info("Variables disponibles para análisis:")
        for i, col in enumerate(descriptors, start=1):
            print(f"{i}: {col}")
        choice = prompt_input(
            message=f"Marque del 1 al {len(descriptors)} para remover, o 0 para continuar",
            default=None,
            validator=lambda x: x.isdigit() and 0 <= int(x) <= len(descriptors),
            error_msg="Debe ingresar un número válido"
        )
        num = int(choice)
        if num == 0:
            break
        removed = descriptors.pop(num-1)
        logger.info(f"Se removió la columna {num}: {removed}")
        if not descriptors:
            logger.error("No quedan variables para analizar. Fin del programa.")
            exit()
    return descriptors

# ----------------------------------------
# Función: build_model_dict
# Descripción: Genera un diccionario con los valores de cada modelo para los descriptores seleccionados.
# ----------------------------------------
def build_model_dict(df, model_col, descriptors):
    """
    Construye un diccionario donde la clave es el nombre del modelo y el valor
    es otro diccionario con pares descriptor:valor.
    """
    model_dict = {}
    for _, row in df.iterrows():
        model = row[model_col]
        values = {col: row[col] for col in descriptors}
        model_dict[model] = values
    logger.info(f"Diccionario de modelos creado con {len(model_dict)} entradas.")
    return model_dict

# ----------------------------------------
# Función: write_dict_file
# Descripción: Guarda el diccionario de modelos en un archivo JSON dentro de la carpeta especificada.
# ----------------------------------------
def write_dict_file(model_dict, folder_path):
    """
    Crea la carpeta si no existe y escribe el diccionario en 'organidez_dict.txt'.
    """
    os.makedirs(folder_path, exist_ok=True)
    txt_path = os.path.join(folder_path, 'organidez_dict.txt')
    with open(txt_path, 'w', encoding='utf-8') as f:
        json.dump(model_dict, f, ensure_ascii=False, indent=2)
    logger.info(f"Diccionario guardado en: {txt_path}")
    return txt_path

# ----------------------------------------
# Función: prompt_top_value
# Descripción: Solicita al usuario el valor del top (número de modelos a considerar).
# ----------------------------------------
def prompt_top_value(max_val):
    """
    Pide al usuario un número entero entre 1 y max_val (inclusive).
    """
    return int(prompt_input(
        message=f"Ingrese un valor de top (1 a {max_val})",
        default=None,
        validator=lambda x: x.isdigit() and 1 <= int(x) <= max_val,
        error_msg="Debe ingresar un número entero dentro del rango"
    ))

# ----------------------------------------
# Función: create_sort_sheet
# Descripción: Crea la hoja 'Sort' en el Excel con los modelos ordenados por descriptor.
# ----------------------------------------
def create_sort_sheet(writer, model_dict, descriptors, top_value):
    """
    Genera la hoja de cálculo 'Sort' con:
    - Columna '#top': ranking del 1 al top_value.
    - Una columna por cada descriptor, con los modelos ordenados.
      R2 de mayor a menor, otros de menor a mayor.
    Excluye modelos con valor NaN para cada descriptor.
    """
    sort_data = {'#top': list(range(1, top_value+1))}
    for col in descriptors:
        valid_items = [(m, vals[col]) for m, vals in model_dict.items() if pd.notna(vals[col])]
        rev = (col == 'R2')
        sorted_models = sorted(valid_items, key=lambda x: x[1], reverse=rev)
        sort_data[col] = [m for m, _ in sorted_models[:top_value]]
        logger.info(f"Hoja 'Sort': procesado descriptor '{col}'.")
    df_sort = pd.DataFrame(sort_data)
    df_sort.to_excel(writer, sheet_name='Sort', index=False)
    return df_sort

# ----------------------------------------
# Función: create_ranking_sheet
# Descripción: Crea la hoja 'Ranking' con tablas de ranking global y de componentes.
# ----------------------------------------
def create_ranking_sheet(writer, df_sort):
    """
    Genera en la hoja 'Ranking':
    1) Ranking global de modelos (#Rank, Model, N°Success).
    2) Tres tablas para Normalization, Reduction y Algorithm,
       con conteo de apariciones en el top.
    """
    from collections import Counter

    cnt = Counter()
    for col in df_sort.columns:
        if col == '#top': continue
        cnt.update([m for m in df_sort[col] if pd.notna(m)])
    models, counts = zip(*cnt.most_common())
    df_rank = pd.DataFrame({'#Rank': list(range(1, len(models)+1)), 'Model': models, 'N°Success': counts})
    logger.info("Hoja 'Ranking': tabla global creada.")

    norm_cnt, red_cnt, alg_cnt = Counter(), Counter(), Counter()
    for col in df_sort.columns:
        if col == '#top': continue
        for model in df_sort[col]:
            parts = model.split('_')
            if len(parts) == 3:
                norm_cnt[parts[0]] += 1
                red_cnt[parts[1]] += 1
                alg_cnt[parts[2]] += 1
    def counter_to_df(cnt):
        items = cnt.most_common()
        rows = [(i+1, nom, val) for i, (nom, val) in enumerate(items)]
        return pd.DataFrame(rows, columns=['#Rank', 'Name', 'N°Success'])

    df_norm, df_red, df_alg = counter_to_df(norm_cnt), counter_to_df(red_cnt), counter_to_df(alg_cnt)
    logger.info("Hoja 'Ranking': tablas de componentes creadas.")

    df_rank.to_excel(writer, sheet_name='Ranking', index=False, startcol=0)
    df_norm.to_excel(writer, sheet_name='Ranking', index=False, startcol=4)
    df_red.to_excel(writer, sheet_name='Ranking', index=False, startcol=8)
    df_alg.to_excel(writer, sheet_name='Ranking', index=False, startcol=12)

# ----------------------------------------
# Función: ajustar_ancho_columnas_excel
# ----------------------------------------
def ajustar_ancho_columnas_excel(path_excel):
    """
    Ajusta el ancho de las columnas automáticamente para cada hoja del archivo Excel.
    """
    wb = load_workbook(path_excel)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            column_letter = get_column_letter(column)
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(path_excel)
    logger.info(f"Ajuste de anchos de columnas aplicado: {path_excel}")

# ----------------------------------------
# Bloque principal
# ----------------------------------------
def main():
    """
    1) Leer ruta de Excel.
    2) Filtrar descriptores.
    3) Generar y guardar diccionario de modelos.
    4) Crear Excel con hojas 'Sort' y 'Ranking'.
    """
    logger.info("Iniciando programa Organizer.")

    # 1. Solicita la ruta del archivo Excel.
    excel_path = prompt_excel_path()

    # 2. Establece el directorio base y la carpeta de salida Organizer_Folder.
    base_dir = os.path.dirname(os.path.abspath(excel_path)) or '.'
    organizer_folder = os.path.join(base_dir, 'Organizer_Folder')

    # 3. Carga el archivo Excel como DataFrame.
    df = pd.read_excel(excel_path, sheet_name=0)

    # 4. Verifica que tenga al menos 3 columnas (entry, modelo, descriptor).
    all_cols = df.columns.tolist()
    if len(all_cols)<3:
        logger.error("Se requieren al menos 3 columnas (entry, modelo, descriptor).")
        return
    
    # 5. Identifica la columna de modelos y lista los descriptores.
    model_col = all_cols[1]
    descriptors = all_cols[2:].copy()

    # 6. Solicita al usuario eliminar descriptores no deseados.
    descriptors = prompt_remove_columns(descriptors)
    if not descriptors:
        return

    # 7. Construye un diccionario con los modelos y valores por descriptor.
    model_dict = build_model_dict(df, model_col, descriptors)

    # 8. Guarda el diccionario en un archivo JSON dentro de Organizer_Folder.
    write_dict_file(model_dict, organizer_folder)

    # 9. Solicita el valor "top" para generar los rankings.
    top_value = prompt_top_value(len(model_dict))

    # 10. Crea el archivo Excel final con hojas "Sort" y "Ranking".
    out_excel = os.path.join(organizer_folder, f'Organizer_top{top_value}.xlsx')
    with pd.ExcelWriter(out_excel, engine='openpyxl') as writer:
        df_sort = create_sort_sheet(writer, model_dict, descriptors, top_value)
        create_ranking_sheet(writer, df_sort)

    # 11. Ajusta automáticamente el ancho de las columnas del Excel generado.
    ajustar_ancho_columnas_excel(out_excel)

    logger.info(f"Archivo Organizer_top{top_value}.xlsx generado en: {out_excel}")

if __name__ == '__main__':
    main()
