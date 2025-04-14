import os
import glob
import logging
import statistics
from openpyxl import load_workbook, Workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# --- Función para leer una hoja "Resultados" y seleccionar el mejor calculado --- #
def get_best_computed_from_resultados(file_path):
    """Carga el Excel de resultados (Resultados_batch_n.xlsx) y
    compara cada columna computada (columna 2 en adelante) con la columna 1 (medido).
    Retorna una tupla: (measured, best_computed, best_label)"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Resultados"]
    measured = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row+1)]
    best_error = None
    best_values = None
    best_label = None
    for col in range(2, ws.max_column+1):
        computed = [ws.cell(row=i, column=col).value for i in range(2, ws.max_row+1)]
        # Calculamos error medio absoluto
        errors = [abs(m - c) for m, c in zip(measured, computed)]
        avg_error = statistics.mean(errors)
        if best_error is None or avg_error < best_error:
            best_error = avg_error
            best_values = computed
            best_label = ws.cell(row=1, column=col).value
    return measured, best_values, best_label

# --- Nivel 1: Procesar Resultados_batch en grupos (sub-batches) --- #
def process_sub_batches(resultados_dir, sub_batch_size, output_dir):
    """
    Busca en resultados_dir todos los archivos "Resultados_batch_*.xlsx",
    los agrupa en sub-batches de tamaño sub_batch_size,
    y para cada grupo crea un Excel que en:
      - Columna 1: valores medidos (se asumen iguales para todas las hojas)
      - Columnas 2..(sub_batch_size+1): la columna ganadora (best computed) de cada archivo en ese grupo.
    Retorna una lista de nombres de archivo de estos sub-batch.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    archivos = sorted(glob.glob(os.path.join(resultados_dir, "Resultados_batch_*.xlsx")))
    sub_batch_files = []
    for idx, grupo in enumerate([archivos[i:i+sub_batch_size] for i in range(0, len(archivos), sub_batch_size)], start=1):
        winners = []
        labels = []
        measured = None
        for file in grupo:
            m, best, label = get_best_computed_from_resultados(file)
            winners.append(best)
            labels.append(label)
            if measured is None:
                measured = m
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparaciones"
        # Escribimos la columna de medido
        ws.cell(row=1, column=1, value="ΔΔG‡ medido")
        for row, val in enumerate(measured, start=2):
            ws.cell(row=row, column=1, value=val)
        # Escribimos cada ganador en una columna separada (a partir de la columna 2)
        for j, (winner, lab) in enumerate(zip(winners, labels), start=2):
            ws.cell(row=1, column=j, value=lab)
            for row, val in enumerate(winner, start=2):
                ws.cell(row=row, column=j, value=val)
        output_file = os.path.join(output_dir, f"Comparaciones_sub_batch_{idx}.xlsx")
        wb.save(output_file)
        logging.info(f"Sub-batch {idx} guardado: {output_file}")
        sub_batch_files.append(output_file)
    return sub_batch_files

# --- Nivel 2: Agrupar los archivos Comparaciones_sub_batch y seleccionar el ganador por grupo --- #
def process_final_batches(sub_batch_files, final_batch_size, output_dir):
    """
    Toma los archivos de Comparaciones_sub_batch, los agrupa en grupos de final_batch_size,
    y para cada grupo selecciona, para cada archivo, la columna (entre las columnas de ganador)
    que esté más cerca del valor medido.
    Luego genera un Excel para cada grupo, donde:
      - Columna 1: ΔΔG‡ medido (fijo)
      - Columnas 2..: la columna ganadora resultante de cada Comparaciones_sub_batch.
    Retorna la lista de archivos finales.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    final_files = []
    for idx, grupo in enumerate([sub_batch_files[i:i+final_batch_size] for i in range(0, len(sub_batch_files), final_batch_size)], start=1):
        group_winners = []
        labels = []
        measured = None
        for file in grupo:
            wb = load_workbook(file, data_only=True)
            ws = wb["Comparaciones"]
            measured = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row+1)]
            best_error = None
            best_values = None
            best_label = None
            # En estos archivos, las columnas 2 en adelante son los ganadores del batch.
            for col in range(2, ws.max_column+1):
                computed = [ws.cell(row=i, column=col).value for i in range(2, ws.max_row+1)]
                errors = [abs(m - c) for m, c in zip(measured, computed)]
                avg_error = statistics.mean(errors)
                if best_error is None or avg_error < best_error:
                    best_error = avg_error
                    best_values = computed
                    best_label = ws.cell(row=1, column=col).value
            group_winners.append(best_values)
            labels.append(best_label)
        # Crear Excel final para este grupo:
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = "Final Comparaciones"
        ws_new.cell(row=1, column=1, value="ΔΔG‡ medido")
        for row, val in enumerate(measured, start=2):
            ws_new.cell(row=row, column=1, value=val)
        for j, (winner, lab) in enumerate(zip(group_winners, labels), start=2):
            ws_new.cell(row=1, column=j, value=lab)
            for row, val in enumerate(winner, start=2):
                ws_new.cell(row=row, column=j, value=val)
        output_file = os.path.join(output_dir, f"Final_batch_{idx}.xlsx")
        wb_new.save(output_file)
        logging.info(f"Final batch {idx} guardado: {output_file}")
        final_files.append(output_file)
    return final_files

# --- Función principal para comparación ---
def main_comparacion():
    # Directorios
    resultados_dir = "C:\\Linux\\resultados"  # donde están los Resultados_batch_*.xlsx
    # Procesar nivel 1: crear archivos Comparaciones_sub_batch
    sub_batch_files = process_sub_batches(resultados_dir, sub_batch_size=100, output_dir=f"{resultados_dir}\\Comparaciones")
    # Procesar nivel 2: agrupar los archivos comparaciones en grupos de 100 y seleccionar ganadores finales
    final_files = process_final_batches(sub_batch_files, final_batch_size=100, output_dir=f"{resultados_dir}\\Final")
    print("Archivos finales generados:")
    for f in final_files:
        print(f)

if __name__ == "__main__":
    main_comparacion()
