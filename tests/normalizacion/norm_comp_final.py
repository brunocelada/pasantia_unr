import os
import openpyxl
from openpyxl import Workbook
import numpy as np

def procesar_final_batches(final_batches_dir, sub_final_dir, sub_batch_size=100):
    os.makedirs(sub_final_dir, exist_ok=True)

    # Buscar todos los Final_batch ordenados
    final_files = sorted([f for f in os.listdir(final_batches_dir) if f.startswith("Final_batch_") and f.endswith(".xlsx")])

    datos_medidos = None
    resultados = []

    for idx, file in enumerate(final_files):
        path = os.path.join(final_batches_dir, file)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # Leer datos de la primera hoja
        columnas = []
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            columnas.append([cell.value for cell in col])

        if datos_medidos is None:
            datos_medidos = columnas[0]  # guardar ΔΔG‡ medido solo una vez

        mejor_columna = encontrar_columna_mas_cercana(datos_medidos, columnas[1:])
        resultados.append(mejor_columna)

    # Agrupar en sub_batchs
    sub_batches = [resultados[i:i+sub_batch_size] for i in range(0, len(resultados), sub_batch_size)]

    for sub_batch_idx, sub_batch in enumerate(sub_batches, start=1):
        wb_out = Workbook()
        ws_out = wb_out.active

        # Escribir la columna de ΔΔG‡ medido
        for fila, valor in enumerate(datos_medidos, start=1):
            ws_out.cell(row=fila, column=1).value = valor

        # Escribir cada columna de resultados
        for col_idx, datos in enumerate(sub_batch, start=2):
            for fila, valor in enumerate(datos, start=1):
                ws_out.cell(row=fila, column=col_idx).value = valor

        output_file = os.path.join(sub_final_dir, f"Sub_Final_batch_{sub_batch_idx}.xlsx")
        wb_out.save(output_file)

def encontrar_columna_mas_cercana(ref_col, candidate_columns):
    ref_array = np.array(ref_col, dtype=float)
    best_col = None
    best_dist = float("inf")

    for col in candidate_columns:
        col_array = np.array(col, dtype=float)
        dist = np.linalg.norm(ref_array - col_array)  # distancia Euclidea
        if dist < best_dist:
            best_dist = dist
            best_col = col

    return best_col

def generar_top5_en_excel(sub_final_dir):
    for file in os.listdir(sub_final_dir):
        if file.startswith("Sub_Final_batch_") and file.endswith(".xlsx"):
            path = os.path.join(sub_final_dir, file)
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            datos_medidos = [cell.value for cell in ws['A']]
            todas_las_columnas = []

            for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
                todas_las_columnas.append([cell.value for cell in col])

            distancias = []
            ref_array = np.array(datos_medidos, dtype=float)

            for idx, columna in enumerate(todas_las_columnas):
                col_array = np.array(columna, dtype=float)
                distancia = np.linalg.norm(ref_array - col_array)
                distancias.append((idx + 2, distancia))  # +2 por offset de columna real

            # Top 5 más cercanos
            top5 = sorted(distancias, key=lambda x: x[1])[:5]

            # Crear nueva hoja
            ws_top = wb.create_sheet("TOP_5_RESULTADOS")
            ws_top.append(["Columna", "Distancia"])

            for col, dist in top5:
                ws_top.append([col, dist])

            wb.save(path)

# --- Función principal para comparación ---
def main_comparacion():
    final_batches_dir = "C:\\Linux\\resultados\\Final"
    sub_final_dir = os.path.join(final_batches_dir, "Sub_Finales")

    procesar_final_batches(final_batches_dir, sub_final_dir)

    # Una vez generado, saca el top5 de cada Sub_Final_batch
    generar_top5_en_excel(sub_final_dir)


if __name__ == "__main__":
    main_comparacion()
