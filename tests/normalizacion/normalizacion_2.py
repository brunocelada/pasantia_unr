import os
import subprocess
import sys
import logging
import math
import statistics
import itertools
from openpyxl import Workbook

# Verificar e instalar openpyxl si no está instalado
try:
    import openpyxl
except ImportError:
    print("openpyxl no está instalado. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

logging.basicConfig(filename="registros/script.log", level=logging.INFO, encoding="utf-8",
                    format="%(asctime)s - %(levelname)s - %(module)s:%(lineno)d - %(message)s")

# --- Normalizaciones ---
def min_max_scaling(data):
    min_val = min(data)
    max_val = max(data)
    return [(x - min_val) / (max_val - min_val) if max_val != min_val else 0.0 for x in data]

def z_score(data):
    if len(data) == 0:
        return data
    mean_val = sum(data) / len(data)
    std_val = (sum((x - mean_val)**2 for x in data) / len(data))**0.5
    if std_val == 0:
        logging.info("z_score: desviación estándar 0 detectada; se devuelve lista de ceros.")
        return [0 for _ in data]
    return [(x - mean_val) / std_val for x in data]

def max_abs_scaling(data):
    max_abs = max(abs(x) for x in data)
    return [x / max_abs if max_abs != 0 else 0.0 for x in data]

def robust_scaling(data):
    mediana = statistics.median(data)
    q1 = statistics.quantiles(data, n=4)[0]
    q3 = statistics.quantiles(data, n=4)[2]
    iqr = q3 - q1
    return [(x - mediana) / iqr if iqr != 0 else 0.0 for x in data]

def log_scaling(data):
    return [math.log1p(x) if x > -1 else 0.0 for x in data]

def l2_norm(data):
    norm = math.sqrt(sum(x ** 2 for x in data))
    return [x / norm if norm != 0 else 0.0 for x in data]

def l1_norm(data):
    norm = sum(abs(x) for x in data)
    return [x / norm if norm != 0 else 0.0 for x in data]

normalizations_available = {
    "minmax": min_max_scaling,
    "zscore": z_score,
    "maxabs": max_abs_scaling,
    "robust": robust_scaling,
    "log": log_scaling,
    "l2": l2_norm,
    "l1": l1_norm
}

# --- Títulos ---
# El archivo de texto no tiene encabezados. Definimos:
# Columna 1: identificador; columnas 2 a 9: features para la ecuación; columna 10: ΔΔG‡ medido (valor medido original)
titles = ["entry (global)",
          "Polarizability",
          "NBO (X)",
          "B5 (avg)",
          "i(N-H)-Achiral",
          "NBO (N)-Chiral",
          "NBO (H)-Chiral",
          "B1 (N-Achiral)",
          "PEOE1",
          "ΔΔG‡ medido"]

# --- Leer archivo y crear diccionario ---
def crear_dict(carpeta_base, archivo_txt):
    os.chdir(carpeta_base)
    try:
        with open(archivo_txt, 'r', encoding='utf-8') as f:
            logging.info(f"Archivo txt abierto: {archivo_txt}")
            datos_dict = {}
            for linea in f:
                if not linea.strip():
                    continue
                partes = linea.strip().split(';')
                if len(partes) != len(titles):
                    logging.warning(f"Línea mal formateada: {linea.strip()}")
                    continue
                key = partes[0]
                try:
                    values = [float(valor) for valor in partes[1:]]
                except ValueError:
                    logging.warning(f"Error al convertir a float en línea: {linea.strip()}")
                    continue
                features_dict = dict(zip(titles[1:9], values[:8]))
                # La última columna (medido) se guarda sin normalizar
                features_dict[titles[9]] = values[8]
                datos_dict[key] = features_dict
            return datos_dict
    except Exception as e:
        logging.error(f"Error encontrando archivo {archivo_txt}: {e}")

# --- Función para calcular la ecuación ---
def calcular_ΔΔG_calculado(features):
    return (1.4 
            - 0.06 * features["PEOE1"]
            + 0.57 * features["NBO (X)"]
            + 0.23 * features["B5 (avg)"]
            + 0.11 * features["Polarizability"]
            - 0.31 * features["i(N-H)-Achiral"]
            + 0.11 * features["NBO (N)-Chiral"]
            - 0.35 * features["NBO (H)-Chiral"]
            - 0.25 * features["B1 (N-Achiral)"])

# --- Función para aplicar UNA normalización o nada (sin cadenas) ---
def aplicar_normalizacion_simple(data, option):
    # option es una tupla; si está vacía, devuelve data sin cambios;
    # si contiene un único método, lo aplica.
    if not option:
        return data[:]
    method = option[0]
    return normalizations_available[method](data)

# --- Generador de combinaciones simples ---
def generate_full_combinations_generator_simple():
    # Cada columna tiene 8 opciones: identidad () y una tupla con cada uno de los métodos.
    opciones = [()] + [(m,) for m in normalizations_available.keys()]
    columnas = titles[1:9]  # Las 8 columnas de features
    for combo in itertools.product(opciones, repeat=8):
        yield dict(zip(columnas, combo))

# --- Calcular resultados para cada combinación ---
def calcular_resultados_combinaciones_general_simple(data_dict, combinaciones):
    keys_ordenados = sorted(data_dict.keys(), key=lambda x: float(x))
    feature_cols = titles[1:9]
    # Extraer los valores originales para cada feature, en el mismo orden
    columnas_originales = {col: [data_dict[k][col] for k in keys_ordenados] for col in feature_cols}
    
    resultados = {}
    for combo in combinaciones:
        # Crear descripción: por ejemplo, "Polarizability: identidad; NBO (X): minmax; ..."
        combo_descr = "; ".join(f"{col}: {'-'.join(option) if option else 'identidad'}" 
                                 for col, option in combo.items())
        columnas_norm = {}
        for col in feature_cols:
            option = combo[col]
            columnas_norm[col] = aplicar_normalizacion_simple(columnas_originales[col], option)
        calculados = []
        for i in range(len(keys_ordenados)):
            features = {col: columnas_norm[col][i] for col in feature_cols}
            calculado = calcular_ΔΔG_calculado(features)
            calculados.append(calculado)
        resultados[combo_descr] = calculados
    return resultados, keys_ordenados

# --- Guardar resultados en Excel ---
def guardar_resultados_en_excel(resultados, data_dict, keys_ordenados, nombre_archivo="Resultados.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.cell(row=1, column=1, value="ΔΔG‡ medido")
    combos = list(resultados.keys())
    for i, combo_descr in enumerate(combos, start=2):
        ws.cell(row=1, column=i, value=combo_descr)
    for idx, key in enumerate(keys_ordenados, start=2):
        ws.cell(row=idx, column=1, value=data_dict[key][titles[9]])
        for j, combo_descr in enumerate(combos, start=2):
            ws.cell(row=idx, column=j, value=resultados[combo_descr][idx-2])
    if not os.path.exists("resultados"):
        os.makedirs("resultados")
    ruta = os.path.join("resultados", nombre_archivo)
    wb.save(ruta)
    print(f"\n✅ Archivo guardado como: {ruta}")

# --- Procesar combinaciones en batches ---
def process_batches_simple(data_dict, batch_size=100):
    gen = generate_full_combinations_generator_simple()
    batch_number = 0
    while True:
        batch = list(itertools.islice(gen, batch_size))
        if not batch:
            break
        resultados, keys_ordenados = calcular_resultados_combinaciones_general_simple(data_dict, batch)
        archivo = f"Resultados_batch_{batch_number}.xlsx"
        guardar_resultados_en_excel(resultados, data_dict, keys_ordenados, nombre_archivo=archivo)
        logging.info(f"Batch {batch_number} procesado y guardado en {archivo}.")
        print(f"Batch {batch_number} procesado y guardado en {archivo}.")
        batch_number += 1

# --- Función principal ---
def main():
    logging.info("\n\n-------NUEVO CÁLCULO SIMPLE (sin cadenas)-------\n")
    carpeta_base = "C:\\Linux"
    archivo_txt = "data_norm.txt"
    
    data_dict = crear_dict(carpeta_base, archivo_txt)
    if not data_dict:
        print("No se pudo cargar el archivo de datos.")
        return
    
    # Generar y procesar todas las combinaciones posibles simples.
    # Total combinaciones = 8^8 = 16.777.216.
    # El programa procesará en batches de 'batch_size' combinaciones.
    process_batches_simple(data_dict, batch_size=100)
    
    print("\nProceso finalizado.\n")

if __name__ == "__main__":
    main()
