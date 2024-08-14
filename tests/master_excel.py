import logging
import sys
import subprocess
import csv
import os

# Configuración de logging
logging.basicConfig(filename="registros/script.log", level=logging.INFO, encoding="utf-8",
                    format="%(asctime)s - %(levelname)s - %(module)s:%(lineno)d - %(message)s")

# verificar e instalar openpyxl y pandas si no están instalados
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    print("openpyxl ya está instalado.")
except ImportError:
    print("openpyxl no está instalado. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
try:
    import pandas as pd
    print("pandas ya está instalado.")
except ImportError:
    print("pandas no está instalado. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
    import pandas as pd


def get_excel_files(directory):
    return [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith(('.xlsx', '.csv'))]

def read_csv_to_dataframe(csv_file):
    with open(csv_file, 'r') as file:
        try:
            df = pd.read_csv(file)
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
    return df

def auto_adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        adjusted_width = max_length + 2
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width

def process_excel_files(directory, master_file_name):
    master_workbook = Workbook()
    master_workbook.remove(master_workbook.active)  # Remove the default sheet

    excel_files = get_excel_files(directory)
    
    for file in excel_files:
        logging.info(f"Processing file: {file}")
        file_name = os.path.splitext(os.path.basename(file))[0]
        
        if file.endswith('.xlsx'):
            workbook = load_workbook(file)
            sheet_number = 1
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                if ws.max_row > 1:
                    new_sheet_name = f"{file_name}-{sheet_number}" if sheet_number > 1 else file_name
                    master_sheet = master_workbook.create_sheet(title=new_sheet_name)
                    for row in ws.iter_rows(values_only=True):
                        master_sheet.append(row)
                    auto_adjust_column_width(master_sheet)
                    logging.info(f"Added sheet {new_sheet_name} from {file}")
                    sheet_number += 1
                    
                else:
                    logging.info(f"Skipped empty sheet {sheet} in {file}")

        elif file.endswith('.csv'):
            df = read_csv_to_dataframe(file)
            if not df.empty:
                master_sheet = master_workbook.create_sheet(title=file_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    master_sheet.append(r)
                auto_adjust_column_width(master_sheet)
                logging.info(f"Added sheet {file_name} from {file}")
            else:
                logging.info(f"Skipped empty CSV file {file}")

    master_file_path = os.path.join(directory, master_file_name)
    master_workbook.save(master_file_path)
    logging.info(f"Master Excel file '{master_file_path}' created successfully.")

def main():
    # Registrar un nuevo lanzamiento del script
    logging.info("\n\n-------NEW MASTER EXCEL-------\n")

    directory = "C:\\Linux"
    master_file_name = input("Enter the name for the master Excel file: ")
    process_excel_files(directory, master_file_name + ".xlsx")

if __name__ == "__main__":
    main()
